from fastapi import APIRouter, HTTPException, Query
from typing import Optional, List
from pydantic import BaseModel, EmailStr
from datetime import datetime
from app.database import get_db_connection
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from io import BytesIO
import hashlib
import binascii
import secrets

router = APIRouter(prefix="/api/users", tags=["users"])

# Pydantic models for request/response
class UserCreate(BaseModel):
    emp_id: int
    name: str
    email: EmailStr
    department: Optional[str] = None
    designation: Optional[str] = None
    hod: Optional[str] = None
    supervisor: Optional[str] = None
    password: str

class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[EmailStr] = None
    department: Optional[str] = None
    designation: Optional[str] = None
    hod: Optional[str] = None
    supervisor: Optional[str] = None

class UserResponse(BaseModel):
    id: int
    emp_id: int
    name: str
    email: str
    department: Optional[str]
    designation: Optional[str]
    hod: Optional[str]
    supervisor: Optional[str]
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True

def hash_password(password: str, salt: Optional[str] = None) -> tuple:
    """
    PBKDF2-HMAC-SHA256 hashing to match auth_router.py.
    Returns (password_hash, salt).
    """
    if salt is None:
        salt = binascii.hexlify(os.urandom(16)).decode()
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 100_000)
    return binascii.hexlify(dk).decode(), salt

@router.get("/", response_model=List[UserResponse])
def get_all_users():
    """Get all users"""
    try:
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("CALL sp_GetAllUsers()")
                users = cursor.fetchall()
        finally:
            connection.close()

        if not users:
            return []

        return users
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export-to-excel")
def export_users_to_excel():
    """Export all users to Excel file"""
    try:
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("CALL sp_GetAllUsers()")
                users = cursor.fetchall()
        finally:
            connection.close()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"

        # Header row with styling
        headers = ["ID", "Employee ID", "Name", "Email", "Department", "Designation", "HOD", "Supervisor", "Created At", "Updated At"]
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Add data rows
        for user in users:
            ws.append([
                user.get('id'),
                user.get('emp_id'),
                user.get('name'),
                user.get('email'),
                user.get('department', ''),
                user.get('designation', ''),
                user.get('hod', ''),
                user.get('supervisor', ''),
                user.get('created_at'),
                user.get('updated_at')
            ])

        # Adjust column widths
        column_widths = [8, 15, 20, 25, 20, 20, 15, 20, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        # Center align all data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Return file content (binary) and filename in JSON for your frontend to consume
        return {
            "success": True,
            "message": "Users exported to Excel successfully",
            "file": output.getvalue(),
            "filename": f"users_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/{emp_id}", response_model=UserResponse)
def get_user_by_emp_id(emp_id: int):
    """Get user by employee ID"""
    try:
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("CALL sp_GetUserByEmpId(%s)", (emp_id,))
                user = cursor.fetchone()
        finally:
            connection.close()

        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        return user
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/{emp_id}", response_model=UserResponse)
def update_user(emp_id: int, user_data: UserUpdate):
    """Update user by employee ID"""
    try:
        # Check if user exists
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("CALL sp_GetUserByEmpId(%s)", (emp_id,))
                existing_user = cursor.fetchone()
                
                if not existing_user:
                    raise HTTPException(status_code=404, detail="User not found")

                # Prepare values for update (using existing values if new ones are None)
                name = user_data.name if user_data.name is not None else existing_user.get('name')
                email = user_data.email if user_data.email is not None else existing_user.get('email')
                dept = user_data.department if user_data.department is not None else existing_user.get('department')
                desig = user_data.designation if user_data.designation is not None else existing_user.get('designation')
                hod = user_data.hod if user_data.hod is not None else existing_user.get('hod')
                sup = user_data.supervisor if user_data.supervisor is not None else existing_user.get('supervisor')

                cursor.execute(
                    "CALL sp_UpdateUser(%s, %s, %s, %s, %s, %s, %s)",
                    (emp_id, name, email, dept, desig, hod, sup)
                )
                connection.commit()
                updated_user = cursor.fetchone()
        finally:
            connection.close()

        return updated_user
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/{emp_id}")
def delete_user(emp_id: int):
    """Delete user by employee ID"""
    try:
        # Check if user exists
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("CALL sp_GetUserByEmpId(%s)", (emp_id,))
                user = cursor.fetchone()
                
                if not user:
                    raise HTTPException(status_code=404, detail="User not found")

                cursor.execute("CALL sp_DeleteUser(%s)", (emp_id,))
                connection.commit()
        finally:
            connection.close()

        return {
            "success": True,
            "message": f"User '{user['name']}' (emp_id: {emp_id}) deleted successfully"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/create")
def create_user(payload: UserCreate):
    try:
        connection = get_db_connection()
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT id FROM users WHERE email = %s", (payload.email,))
                if cursor.fetchone():
                    raise HTTPException(status_code=409, detail="User already exists")

                pwd_hash, salt = hash_password(payload.password)
                cursor.execute(
                    "CALL sp_CreateUser(%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (payload.emp_id, payload.name, payload.email, payload.department, payload.designation, payload.hod, payload.supervisor, pwd_hash, salt)
                )
                connection.commit()
                user = cursor.fetchone()
        finally:
            connection.close()
        return {"success": True, "message": "User created", "data": user}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

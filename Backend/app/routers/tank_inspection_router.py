# app/routers/tank_inspection_router.py
from fastapi import APIRouter, HTTPException, Depends, status, UploadFile, File, Header
from fastapi.responses import JSONResponse
from fastapi.encoders import jsonable_encoder
from pydantic import BaseModel, Field
from datetime import datetime, date
from typing import List, Optional, Generator, Any
from sqlalchemy import func, text, inspect
from sqlalchemy.orm import Session
import os
import uuid
import logging
import traceback
import jwt  # PyJWT
import pymysql
from pymysql.cursors import DictCursor
from decimal import Decimal
import urllib.parse
import importlib

from app.utils.s3_utils import AWS_S3_BUCKET, to_cdn_url
from app.database import get_db, get_db_connection
from app.routers import to_do_list_router
from app.routers.tank_checkpoints_router import FAULTY_STATUS_IDS
from app.models.tank_certificate import TankCertificate
from app.models.inspection_history_model import InspectionHistory
from app.models.users_model import User

try:
    from PIL import Image
except Exception:
    Image = None

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

router = APIRouter(prefix="/api/tank_inspection_checklist", tags=["tank_inspection"])

UPLOAD_DIR = "uploads"
if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)
IMAGES_ROOT_DIR = os.path.join(UPLOAD_DIR, "tank_images_mobile")
if not os.path.exists(IMAGES_ROOT_DIR):
    os.makedirs(IMAGES_ROOT_DIR, exist_ok=True)

JWT_SECRET = os.getenv("JWT_SECRET", "change_this_in_production")
JWT_ALGORITHM = os.getenv("JWT_ALGORITHM", "HS256")


def _is_blank_or_zero(v):
    """Return True if value is None, empty string, or numeric 0 (or "0")."""
    if v is None:
        return True
    if isinstance(v, str) and v.strip() == "":
        return True
    try:
        return int(v) == 0
    except Exception:
        return False


# Response helpers (uniform envelope)
from fastapi.encoders import jsonable_encoder

def success_resp(message: str, data: Any = None, status_code: int = 200):
    return JSONResponse(
        status_code=status_code,
        content={
            "success": True,
            "message": message,
            "data": jsonable_encoder(data) if data is not None else {}
        },
    )

def error_resp(message: str, status_code: int = 400):
    return JSONResponse(
        status_code=status_code,
        content={"success": False, "message": message, "data": {}},
    )

# -------------------------------------------------------------------
# AUTH – FIXED AND STABLE
# -------------------------------------------------------------------
def get_current_user(
    authorization: Optional[str] = Header(None, alias="Authorization"),
    db: Session = Depends(get_db)
):
    if not authorization:
        raise HTTPException(status_code=401, detail="Authorization required")

    token = authorization.replace("Bearer", "").strip()

    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

    if not User:
        raise HTTPException(status_code=401, detail="User model not available")

    # 🔥 FIX: Always resolve user from DB
    user = None

    if payload.get("emp_id"):
        user = db.query(User).filter(User.emp_id == int(payload["emp_id"])).first()

    if not user and payload.get("login_name"):
        user = db.query(User).filter(User.login_name == payload["login_name"]).first()

    if not user and payload.get("email"):
        user = db.query(User).filter(User.email == payload["email"]).first()

    if not user and payload.get("sub"):
        user = db.query(User).filter(User.login_name == payload["sub"]).first()

    if not user:
        raise HTTPException(status_code=401, detail="User not found")

    # 🔐 SESSION VALIDATION (CRITICAL)
    session = db.execute(
        text("""
            SELECT 1 FROM login_sessions
            WHERE emp_id = :eid AND still_logged_in = 1
            LIMIT 1
        """),
        {"eid": user.emp_id},
    ).fetchone()

    if not session:
        raise HTTPException(status_code=401, detail="Session expired or logged out")

    return user

# -------------------------------------------------------------------
# EMP_ID RESOLVER (FINAL)
# -------------------------------------------------------------------
def resolve_emp_id(current_user):
    if current_user and hasattr(current_user, "emp_id"):
        return int(current_user.emp_id)
    return None
# -------------------------
# File helpers
# -------------------------
import time
from io import BytesIO
from app.utils.s3_utils import build_s3_key, upload_fileobj_to_s3


def fetch_pi_next_inspection_date(db: Session, tank_number: str):
    try:
        row = db.execute(text("CALL sp_GetNextInspDate(:tank_number)"), {"tank_number": tank_number}).mappings().first()
        if not row:
            return None
        return row.get("next_insp_date")
    except Exception as exc:
        logger.warning("Could not fetch PI next inspection date for tank_number %s: %s", tank_number, exc)
        return None


def generate_report_number(db: Session, inspection_date: datetime, inspection_type_id: Optional[int] = None) -> str:
    date_str = inspection_date.strftime("%d%m%Y")
    
    # Determine Prefix
    prefix_part = "SG-T1" # Default
    if inspection_type_id:
        try:
            it_row = db.execute(text("CALL sp_GetInspectionTypeName(:id)"), {"id": inspection_type_id}).mappings().first()
            if it_row:
                itype_name = it_row.get("inspection_type_name", "").upper()
                normalized_name = itype_name.replace("-", "").replace(" ", "")
                if normalized_name in ["ONHIRE", "OFFHIRE", "CONDITION"]:
                    prefix_part = f"SG-{normalized_name}-T1"
        except Exception as e:
            logger.warning(f"Error fetching inspection type name for report number generation: {e}")

    for attempt in range(3):
        try:
            cnt_row = db.execute(text("CALL sp_GetDailyInspectionCount(:d)"), {"d": inspection_date.date()}).mappings().first()
            count = int(cnt_row.get("cnt", 0)) if cnt_row else 0
        except Exception:
            count = 0

        next_counter = (count or 0) + 1
        report_number = f"{prefix_part}-{date_str}-{next_counter:02d}"

        try:
            existing = db.execute(text("CALL sp_CheckReportNumberExists(:rn)"), {"rn": report_number}).mappings().first()
            if not existing:
                return report_number
        except Exception:
            return report_number

        logger.warning(f"Report number collision for {report_number}, retrying...")

    raise RuntimeError(f"Unable to generate unique report number after retries for date {date_str}")


def fetch_tank_details(db: Session, tank_number: str):
    result = db.execute(text("CALL sp_GetInspectionTankDetails(:tank_number)"), {"tank_number": tank_number}).mappings().first()

    if not result:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Tank details not found for tank_number: {tank_number}",
        )

    return dict(result)


# -------------------------
# Pydantic schemas (updated to use tank_id in create/update)
# -------------------------
class TankInspectionCreate(BaseModel):
    tank_id: int = Field(..., description="tank_details.tank_id (client must send tank_id)")
    status_id: Optional[int] = None
    inspection_type_id: Optional[int] = None
    location_id: Optional[int] = None
    safety_valve_brand_id: Optional[int] = None  # nullable
    safety_valve_model_id: Optional[int] = None  # nullable
    safety_valve_size_id: Optional[int] = None   # nullable
    notes: Optional[str] = None
    operator_id: Optional[int] = None   # manual operator id entered by user (nullable)
    vacuum_reading: Optional[str] = None
    vacuum_uom: Optional[str] = None
    lifter_weight_value: Optional[str] = None
    pi_next_inspection_date: Optional[str] = None

    class Config:
        json_schema_extra = {
            "examples": [
                {
                    "tank_id": 0,
                    "status_id": 0,
                    "inspection_type_id": 0,
                    "vaccum_reading": 0,
                    "vacuum_uom": 0,
                    "lifter_weight_value": 0,
                    "location_id": 0,
                    "safety_valve_model_id": 0,
                    "safety_valve_size_id": 0,
                    "notes": "All checks ok",
                    "operator_id": 0
                }
            ]
        }

class TankInspectionResponse(BaseModel):
    inspection_id: int
    tank_number: str
    report_number: str
    inspection_date: datetime
    status_id: Optional[int] = None
    inspection_type_id: Optional[int] = None
    location_id: Optional[int] = None
    working_pressure: Optional[float] = None
    design_temperature: Optional[str] = None
    frame_type: Optional[str] = None
    cabinet_type: Optional[str] = None
    mfgr: Optional[str] = None
    pi_next_inspection_date: Optional[str] = None
    safety_valve_model_id: Optional[int] = None
    safety_valve_size_id: Optional[int] = None
    notes: Optional[str] = None
    created_by: Optional[str] = None
    operator_id: Optional[int] = None
    emp_id: int     # NOT optional - must be the logged-in user's ID
    ownership: Optional[str] = None
    vacuum_reading: Optional[str] = None
    vacuum_uom: Optional[str] = None
    lifter_weight_value: Optional[str] = None
    created_at: Optional[datetime] = None
    updated_at: Optional[datetime] = None

    class Config:
        from_attributes = True


class TankInspectionUpdate(BaseModel):
    inspection_date: Optional[datetime] = None
    # client can send tank_id if they want to change which tank this inspection refers to (rare) --
    # if provided, code resolves tank_number from tank_id.
    tank_id: Optional[int] = None
    status_id: Optional[int] = None
    inspection_type_id: Optional[int] = None
    location_id: Optional[int] = None
    safety_valve_model_id: Optional[int] = None      # nullable
    safety_valve_size_id: Optional[int] = None       # nullable
    pi_next_inspection_date: Optional[str] = None

    class Config:
        from_attributes = True


# -------------------------
# NEW: GET ALL INSPECTIONS
# -------------------------
@router.get("/list/all")
def get_all_inspections(
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """
    Fetch all inspection records from tank_inspection_details.
    Includes joins for human-readable names.
    """
    try:
        results = db.execute(text("CALL sp_GetAllInspectionDetails()")).mappings().fetchall()
        
        # Convert Decimals and datetimes for JSON, and remove IDs
        def clean_row(r):
            d = dict(r)
            
            # ---------------------------------------------------------
            # Format report_number dynamically to insert inspection_type
            # ---------------------------------------------------------
            report_num = d.get("report_number")
            itype_name = d.get("inspection_type_name")
            if report_num and itype_name and report_num.startswith("SG-T1-"):
                normalized_name = itype_name.upper().replace("-", "").replace(" ", "")
                if normalized_name in ["ONHIRE", "OFFHIRE", "CONDITION"]:
                    d["report_number"] = report_num.replace("SG-T1-", f"SG-{normalized_name}-T1-", 1)
                    
            # Remove IDs as requested to return names "instead of" ids
            for id_key in ["status_id", "inspection_type_id", "location_id"]:
                d.pop(id_key, None)
                
            for k, v in d.items():
                if isinstance(v, Decimal):
                    d[k] = float(v)
                elif isinstance(v, datetime):
                    d[k] = v.isoformat()
                elif isinstance(v, date):
                    d[k] = v.isoformat()
            return d

        data = [clean_row(r) for r in results]
        return success_resp("All inspections fetched", data)

    except Exception as e:
        logger.error(f"Error fetching all inspections: {e}", exc_info=True)
        return error_resp(f"Internal server error: {str(e)}", 500)


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from fastapi.responses import StreamingResponse

# -------------------------
# EXPORT INSPECTIONS TO EXCEL
# -------------------------
@router.get("/export-to-excel")
def export_inspections_to_excel(
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """
    Export all inspection records to an Excel file with detailed information.
    """
    try:
        query = text("""
            SELECT 
                ti.inspection_id,
                ti.report_number,
                ti.tank_number,
                ti.inspection_date,
                ti.mfgr,
                ti.working_pressure,
                ti.design_temperature,
                ti.frame_type,
                ti.cabinet_type,
                ti.pi_next_inspection_date,
                ti.notes,
                ti.vacuum_reading,
                ti.vacuum_uom,
                ti.lifter_weight_value,
                ti.ownership,
                ti.operator_id,
                ti.is_submitted,
                ti.web_submitted,
                ti.created_by,
                ti.created_at,
                ti.updated_by,
                ti.updated_at,
                ps.status_name,
                it.inspection_type_name,
                pl.location_name,
                pm.product_name,
                sb.brand_name AS safety_valve_brand_name,
                sm.model_name AS safety_valve_model_name,
                ss.size_label AS safety_valve_size_name
            FROM tank_inspection_details ti
            JOIN tank_details td ON ti.tank_id = td.tank_id
            LEFT JOIN tank_status ps ON ti.status_id = ps.id
            LEFT JOIN inspection_type it ON ti.inspection_type_id = it.id
            LEFT JOIN location_master pl ON ti.location_id = pl.id
            LEFT JOIN product_master pm ON td.product_id = pm.id
            LEFT JOIN safety_valve_brand sb ON COALESCE(ti.safety_valve_brand_id, td.safety_valve_brand_id) = sb.id
            LEFT JOIN safety_valve_model sm ON ti.safety_valve_model_id = sm.id
            LEFT JOIN safety_valve_size ss ON ti.safety_valve_size_id = ss.id
            ORDER BY ti.inspection_date DESC
        """)
        
        results = db.execute(query).mappings().fetchall()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Inspection Reports"
        
        headers = [
            "ID", "Report No", "Tank No", "Date", "Status", "Inspection Type", "Location", "Product", 
            "Manufacturer", "Working Pressure", "Design Temp", "Frame Type", "Cabinet Type", "Next Inspection", 
            "SV Brand", "SV Model", "SV Size", "Vacuum Reading", "Vacuum UOM", "Lifter Weight", "Ownership", "Operator ID", 
            "Is Submitted", "Web Submitted", "Created By", "Created At", "Updated By", "Updated At"
        ]
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
        for row_num, row_data in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=row_data.get("inspection_id"))
            
            # ---------------------------------------------------------
            # Format report_number dynamically to insert inspection_type
            # ---------------------------------------------------------
            report_num = row_data.get("report_number")
            itype_name = row_data.get("inspection_type_name")
            if report_num and itype_name and report_num.startswith("SG-T1-"):
                normalized_name = itype_name.upper().replace("-", "").replace(" ", "")
                if normalized_name in ["ONHIRE", "OFFHIRE", "CONDITION"]:
                    report_num = report_num.replace("SG-T1-", f"SG-{normalized_name}-T1-", 1)
                    
            ws.cell(row=row_num, column=2, value=report_num)
            ws.cell(row=row_num, column=3, value=row_data.get("tank_number"))
            ws.cell(row=row_num, column=4, value=str(row_data.get("inspection_date")))
            ws.cell(row=row_num, column=5, value=row_data.get("status_name"))
            ws.cell(row=row_num, column=6, value=row_data.get("inspection_type_name"))
            ws.cell(row=row_num, column=7, value=row_data.get("location_name") or "-")
            ws.cell(row=row_num, column=8, value=row_data.get("product_name") or "-")
            ws.cell(row=row_num, column=9, value=row_data.get("mfgr") or "-")
            ws.cell(row=row_num, column=10, value=row_data.get("working_pressure") or "-")
            ws.cell(row=row_num, column=11, value=row_data.get("design_temperature") or "-")
            ws.cell(row=row_num, column=12, value=row_data.get("frame_type") or "-")
            ws.cell(row=row_num, column=13, value=row_data.get("cabinet_type") or "-")
            ws.cell(row=row_num, column=14, value=row_data.get("pi_next_inspection_date") or "-")
            ws.cell(row=row_num, column=15, value=row_data.get("safety_valve_brand_name") or "-")
            ws.cell(row=row_num, column=16, value=row_data.get("safety_valve_model_name") or "-")
            ws.cell(row=row_num, column=17, value=row_data.get("safety_valve_size_name") or "-")
            ws.cell(row=row_num, column=18, value=row_data.get("vacuum_reading") or "-")
            ws.cell(row=row_num, column=19, value=row_data.get("vacuum_uom") or "-")
            ws.cell(row=row_num, column=20, value=row_data.get("lifter_weight_value") or "-")
            ws.cell(row=row_num, column=21, value=row_data.get("ownership") or "-")
            ws.cell(row=row_num, column=22, value=row_data.get("operator_id") or "-")
            
            # Is Submitted status
            is_sub = "SUBMITTED" if row_data.get("is_submitted") == 1 else "DRAFT"
            ws.cell(row=row_num, column=23, value=is_sub)
            
            # Web Submitted status with UPDATED logic
            web_status = ""
            if row_data.get("web_submitted") == 1:
                web_status = "UPDATED" if row_data.get("is_submitted") == 1 else "SUBMITTED"
            ws.cell(row=row_num, column=24, value=web_status)
            
            ws.cell(row=row_num, column=25, value=row_data.get("created_by"))
            ws.cell(row=row_num, column=26, value=str(row_data.get("created_at")))
            ws.cell(row=row_num, column=27, value=row_data.get("updated_by"))
            ws.cell(row=row_num, column=28, value=str(row_data.get("updated_at")))

        # Auto-adjust column width
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            header_val = str(col[0].value)
            max_length = len(header_val)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Inspection_Reports_{timestamp}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Error exporting inspections: {e}", exc_info=True)
        return JSONResponse(status_code=500, content={"success": False, "message": "Failed to export inspections"})


# -------------------------
# Auth helper (kept as before)
# -------------------------
try:
    from app.models.users_model import User
except Exception:
    User = None


def get_current_user(authorization: Optional[str] = Header(None, alias="Authorization"), db: Session = Depends(get_db)):
    if not authorization:
        return None
    auth = authorization.strip()
    token = auth
    if len(auth) >= 6 and auth[:6].lower() == "bearer":
        token_part = auth[6:]
        token = token_part.lstrip(" :\t")
    token = token.strip()
    if not token:
        return None
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    except Exception:
        raise HTTPException(status_code=401, detail="Could not validate credentials")

    if User is None:
        return payload

    user = None
    try:
        if "emp_id" in payload and payload["emp_id"] is not None:
            try:
                user = db.query(User).filter(User.emp_id == int(payload["emp_id"])).first()
            except Exception:
                user = db.query(User).filter(User.emp_id == payload["emp_id"]).first()
        elif "email" in payload and payload["email"]:
            user = db.query(User).filter(User.email == payload["email"]).first()
        elif "sub" in payload and payload["sub"]:
            sub = payload["sub"]
            try:
                user = db.query(User).filter((User.email == sub) | (User.emp_id == int(sub))).first()
            except Exception:
                user = db.query(User).filter((User.email == sub) | (User.emp_id == sub)).first()
    except Exception:
        raise HTTPException(status_code=401, detail="Validation error")

    if not user:
        raise HTTPException(status_code=401, detail="User not found")
        # Enforce session validity (check if logged out)
        
    try:
        session_row = db.execute(
            text("SELECT 1 FROM login_sessions WHERE emp_id = :eid AND still_logged_in = 1 LIMIT 1"),
            {"eid": user.emp_id}
        ).fetchone()
        
        if not session_row:
             # User logged out explicitly
             raise HTTPException(status_code=401, detail="Session invalid or logged out")
    except HTTPException:
        raise
    except Exception as e:
        logger.warning(f"Session check error: {e}")
        # Fail safe to unauthorized if session state cannot be verified
        raise HTTPException(status_code=401, detail="Could not verify session active state")


    return user


@router.get("/auth/debug-token")
def debug_token(authorization: Optional[str] = Header(None, alias="Authorization")):
    if not authorization:
        return error_resp("No Authorization header", 400)
    token = authorization.strip()
    if token.lower().startswith("bearer "):
        token = token.split(" ", 1)[1].strip()
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM], options={"verify_signature": False})
        return success_resp("Decoded token payload (no signature verification)", payload, 200)
    except Exception as e:
        return error_resp(f"Failed to decode token: {e}", 400)


# -------------------------
# Helper: validate operator exists in operators table
# -------------------------
def operator_exists(db: Session, operator_id: int) -> bool:
    try:
        r = db.execute(text("SELECT 1 FROM operators WHERE operator_id = :op LIMIT 1"), {"op": operator_id}).fetchone()
        return bool(r)
    except Exception:
        return False


# -------------------------
# Masters endpoint (kept)
# -------------------------
@router.get("/masters")
def get_all_tank_inspection_masters():
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor(DictCursor) as cursor:
            masters = {
                "tank_statuses": ("tank_status", ["status_id", "status_name", "description", "created_at", "updated_at"]),
                "products": ("product_master", ["product_id", "product_name", "description", "created_at", "updated_at"]),
                "inspection_types": ("inspection_type", ["inspection_type_id", "inspection_type_name", "description", "created_at", "updated_at"]),
                "locations": ("location_master", ["location_id", "location_name", "description", "created_at", "updated_at"]),
                "safety_valve_brands": ("safety_valve_brand", ["id", "brand_name", "description", "created_at", "updated_at"]),
                "safety_valve_models": ("safety_valve_model", ["id", "model_name", "description", "created_at", "updated_at"]),
                "safety_valve_sizes": ("safety_valve_size", ["id", "size_label", "description", "created_at", "updated_at"]),
            }

            out_data = {}

            for key, (table, expected_fields) in masters.items():
                try:
                    cursor.execute(f"SELECT * FROM `{table}` LIMIT 100")
                    sample_rows = cursor.fetchall() or []
                except Exception as ex:
                    logger.warning("Failed to fetch table %s: %s", table, ex, exc_info=True)
                    out_data[key] = []
                    continue

                real_cols = list(sample_rows[0].keys()) if sample_rows else []
                if not real_cols:
                    try:
                        cursor.execute(f"SELECT * FROM `{table}` LIMIT 0")
                        real_cols = [d[0] for d in cursor.description] if cursor.description else []
                    except Exception:
                        real_cols = []

                def pick_real_col_for_expected(ef):
                    if ef.endswith("_id"):
                        if ef in real_cols:
                            return ef
                        if "id" in real_cols:
                            return "id"
                        base = ef[:-3]
                        if f"{base}_id" in real_cols:
                            return f"{base}_id"
                        if f"{base}id" in real_cols:
                            return f"{base}id"
                        return None
                    candidates = [ef]
                    if ef.endswith("_name"):
                        candidates.append(ef.replace("_name", "name"))
                        candidates.append(ef.replace("_name", ""))
                    for c in candidates:
                        if c in real_cols:
                            return c
                    for c in real_cols:
                        if c.lower().endswith(ef.split("_")[-1].lower()):
                            return c
                    return None

                chosen_map = {ef: pick_real_col_for_expected(ef) for ef in expected_fields}
                mapped = []
                for r in sample_rows:
                    out_row = {}
                    for ef in expected_fields:
                        real = chosen_map.get(ef)
                        val = None
                        if real and real in r:
                            val = r.get(real)
                        out_row[ef] = val
                    mapped.append(out_row)
                out_data[key] = jsonable_encoder(mapped)

            return success_resp("Master data fetched successfully", out_data, 200)
    except Exception as e:
        logger.error(f"Error fetching masters: {e}", exc_info=True)
        return error_resp("Error fetching master data", 500)
    finally:
        try:
            if conn:
                conn.close()
        except Exception:
            pass


# Simple validator for tank existence (kept)
def validate_tank_exists(db: Session, tank_number: str):
    result = db.execute(
        text("SELECT 1 FROM tank_header WHERE tank_number = :tank_number"),
        {"tank_number": tank_number},
    ).fetchone()
    if not result:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Tank not existing: {tank_number}")


# -------------------------
# Active tanks endpoint
# -------------------------
@router.get("/active-tanks")
def get_active_tanks(db: Session = Depends(get_db), current_user: Optional[dict] = Depends(get_current_user)):
    try:
        rows = db.execute(text("SELECT tank_id, tank_number FROM tank_details WHERE status = 'active'")).mappings().all()
        data = [dict(r) for r in rows]
        return success_resp("Active tanks fetched", {"active_tanks": jsonable_encoder(data)}, 200)
    except Exception as e:
        logger.error(f"Error fetching active tanks: {e}", exc_info=True)
        return error_resp("Error fetching active tanks", 500)



# -------------------------
# Create Tank Inspection (flat payload with master ids)
# -------------------------
@router.post("/create/tank_inspection", status_code=status.HTTP_201_CREATED)
def create_tank_inspection(
    payload: TankInspectionCreate,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user),
):
    try:
        # --- Resolve tank_number from payload.tank_id ---
        try:
            tn_row = db.execute(
                text("SELECT tank_number FROM tank_details WHERE tank_id = :tid LIMIT 1"),
                {"tid": payload.tank_id},
            ).fetchone()
        except Exception as e:
            logger.error("DB error resolving tank_number: %s", e, exc_info=True)
            return error_resp(f"Tank not found for id: {payload.tank_id}", 404)

        if not tn_row:
            return error_resp(f"Tank not found for id: {payload.tank_id}", 404)

        # Handle row mapping safely
        if hasattr(tn_row, "_mapping"):
            tank_number = tn_row._mapping.get("tank_number")
        elif isinstance(tn_row, dict):
            tank_number = tn_row.get("tank_number")
        else:
            tank_number = tn_row[0]

        # --- Helper: Strictly check if value is a valid ID ---
        def is_valid_id(val):
            if val is None:
                return False
            if isinstance(val, int) and val > 0:
                return True
            if isinstance(val, str) and val.isdigit() and int(val) > 0:
                return True
            return False

        # --- Validate master ids (Only if provided and non-zero) ---
        # Note: During CREATE, all IDs are allowed to be 0 or null
        # Validation happens at SUBMIT time via validation/submit endpoints
        master_checks = [
            ("tank_status", payload.status_id, "status_id"),
            ("inspection_type", payload.inspection_type_id, "inspection_type_id"),
            ("location_id", payload.location_id, "location_id"),
            # Safety valves (Optional)
            ("safety_valve_model", payload.safety_valve_model_id, "safety_valve_model_id"),
            ("safety_valve_size", payload.safety_valve_size_id, "safety_valve_size_id"),
        ]

        # --- Prepare Inspection Data ---
        tank_details = fetch_tank_details(db, tank_number)
        inspection_date = datetime.now()
        
        # Resolve Emp ID (from JWT / session when available; never trust payload)
        emp_id_val = resolve_emp_id(current_user)

        if not emp_id_val:
            return error_resp("Unable to resolve emp_id from token", 401)


        # Check for any existing non-submitted inspection for this tank
        active_inspection = db.execute(
            text("""
                SELECT inspection_id, report_number
                FROM tank_inspection_details 
                WHERE tank_number = :tn 
                  AND (is_submitted IS NULL OR is_submitted = 0)
                  AND (web_submitted IS NULL OR web_submitted = 0)
                  AND (is_reviewed IS NULL OR is_reviewed = 0)
                LIMIT 1
            """),
            {"tn": tank_number}
        ).fetchone()

        if active_inspection:
            report_num = ""
            if hasattr(active_inspection, "_mapping"):
                report_num = active_inspection._mapping.get("report_number")
            elif hasattr(active_inspection, "report_number"):
                report_num = active_inspection.report_number
            else:
                try:
                    report_num = active_inspection[1]
                except Exception:
                    report_num = "Unknown"
                    
            if not report_num:
                report_num = "Unknown"
                
            return error_resp(f"An active, non-submitted inspection (Report No: {report_num}) already exists for tank {tank_number}. Please complete or edit the existing inspection before creating a new one.", 400)

        # Duplicate Check
        existing = db.execute(
            text("SELECT inspection_id FROM tank_inspection_details WHERE tank_number = :tn AND DATE(inspection_date) = :d AND inspection_type_id = :itype LIMIT 1"),
            {"tn": tank_number, "d": inspection_date.date(), "itype": payload.inspection_type_id},
        ).fetchone()
        if existing:
            return error_resp("Inspection already exists", 400)

        # Generate Reports
        report_number = generate_report_number(db, inspection_date, inspection_type_id=payload.inspection_type_id)
        pi_next_date = payload.pi_next_inspection_date if payload.pi_next_inspection_date else fetch_pi_next_inspection_date(db, tank_number)
        ownership_val = tank_details.get("ownership")
        # Sanitize Safety Valve IDs for Insert (Ensure None if invalid)
        svb = payload.safety_valve_brand_id if is_valid_id(payload.safety_valve_brand_id) else None
        svm = payload.safety_valve_model_id if is_valid_id(payload.safety_valve_model_id) else None
        svs = payload.safety_valve_size_id if is_valid_id(payload.safety_valve_size_id) else None

        # --- INSERT ---
        try:
            db.execute(
                text("""
                    INSERT INTO tank_inspection_details
                    (inspection_date, report_number, tank_number, tank_id, status_id, inspection_type_id, location_id,
                     working_pressure, frame_type, design_temperature, cabinet_type, mfgr, pi_next_inspection_date,
                     safety_valve_brand_id, safety_valve_model_id, safety_valve_size_id, notes,
                     vacuum_reading, vacuum_uom, lifter_weight_value,
                     created_by, updated_by,
                     operator_id, emp_id, ownership, is_submitted, created_at, updated_at)
                    VALUES
                    (:inspection_date, :report_number, :tank_number, :tank_id, :status_id, :inspection_type_id, :location_id,
                     :working_pressure, :frame_type, :design_temperature, :cabinet_type, :mfgr, :pi_next_inspection_date,
                     :svb, :svm, :svs, :notes,
                     :vacuum_reading, :vacuum_uom, :lifter_weight_value,
                     :created_by, :updated_by,
                     :operator_id, :emp_id, :ownership, :is_submitted, NOW(), NOW())
                """),
                {
                    "inspection_date": inspection_date,
                    "report_number": report_number,
                    "tank_number": tank_number,
                    "tank_id": payload.tank_id,
                    "status_id": None if payload.status_id in [None, 0, "0", ""] else payload.status_id,
                    "inspection_type_id": payload.inspection_type_id if is_valid_id(payload.inspection_type_id) else None,
                    "location_id": payload.location_id if is_valid_id(payload.location_id) else None,
                    "working_pressure": tank_details.get("working_pressure"),
                    "frame_type": tank_details.get("frame_type"),
                    "design_temperature": tank_details.get("design_temperature"),
                    "cabinet_type": tank_details.get("cabinet_type"),
                    "mfgr": tank_details.get("mfgr"),
                    "pi_next_inspection_date": pi_next_date,
                    "svb": svb, "svm": svm, "svs": svs,
                    "notes": payload.notes,
                    "vacuum_reading": payload.vacuum_reading,
                    "vacuum_uom": payload.vacuum_uom,
                    "lifter_weight_value": payload.lifter_weight_value,
                    # created_by / updated_by MUST come from logged-in user
                    "created_by": emp_id_val,
                    "updated_by": emp_id_val,
                    "operator_id": payload.operator_id,
                    "emp_id": emp_id_val,
                    "ownership": ownership_val,
                    "is_submitted": 0,
                },
            )
            
            # --- SYNCHRONIZE TANK MASTER ---
            if svb is not None:
                db.execute(
                    text("UPDATE tank_details SET safety_valve_brand_id = :svb WHERE tank_number = :tn"),
                    {"svb": svb, "tn": tank_number}
                )
            
            db.commit()
        except Exception as e:
            db.rollback()
            logger.error("Failed to create tank inspection record: %s", e, exc_info=True)
            return error_resp(f"Internal server error: {e}", 500)

        # --- Return Created Record ---
        new_row = db.execute(text("SELECT * FROM tank_inspection_details WHERE report_number = :rn"), {"rn": report_number}).fetchone()
        
        # Convert row to dict safely
        if hasattr(new_row, "_mapping"):
            out = dict(new_row._mapping)
        elif isinstance(new_row, dict):
            out = new_row
        else:
            out = dict(zip(new_row.keys(), new_row))

        return success_resp("Inspection created successfully", out, 201)

    except Exception as e:
        logger.error(f"Error creating tank inspection: {e}", exc_info=True)
        return error_resp(f"Internal server error: {e}", 500)


# -------------------------
# Update tank_inspection_details (PUT)
# -------------------------
class TankInspectionUpdateModel(BaseModel):
    inspection_date: Optional[datetime] = None
    tank_id: Optional[int] = None
    status_id: Optional[int] = None
    inspection_type_id: Optional[int] = None
    location_id: Optional[int] = None
    safety_valve_brand_id: Optional[int] = None      # nullable
    safety_valve_model_id: Optional[int] = None      # nullable
    safety_valve_size_id: Optional[int] = None       # nullable
    vacuum_reading: Optional[str] = None
    vacuum_uom: Optional[str] = None
    lifter_weight_value: Optional[str] = None
    pi_next_inspection_date: Optional[str] = None

    class Config:
        from_attributes = True


@router.put("/update/tank_inspection_details/{inspection_id}")
def update_tank_inspection_details(
    inspection_id: int, 
    payload: TankInspectionUpdateModel, 
    db: Session = Depends(get_db), 
    current_user: Optional[dict] = Depends(get_current_user)
):
    try:
        # 1. Check if inspection exists and role permissions
        row = db.execute(text("SELECT is_submitted, web_submitted, tank_id, tank_number FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        if not row:
            return error_resp("Inspection not found", 404)

        is_submitted = int(row[0])
        web_submitted = int(row[1])
        current_tank_id = int(row[2])
        current_tank_number = str(row[3])
        role_id = current_user.role_id
        if (is_submitted == 1 or web_submitted == 1) and role_id == 2:
            return error_resp("Cannot edit submitted inspection", 403)

        params = {"id": inspection_id}
        updates = []

        # Helper: Strictly check if value is a valid ID (int > 0)
        def is_valid_id(val):
            if val is None: return False
            if isinstance(val, int) and val > 0: return True
            if isinstance(val, str) and val.isdigit() and int(val) > 0: return True
            return False

        # Helper: Get set fields safely (Pydantic v1/v2 compat)
        try:
            # Try Pydantic v2
            update_data = payload.model_dump(exclude_unset=True)
        except AttributeError:
            # Fallback to Pydantic v1
            update_data = payload.dict(exclude_unset=True)

        # --- Handle Special Fields (operator_id, emp_id, tank_id) ---
        
        # Operator ID (Optional)
        if "operator_id" in update_data:
            op_id = update_data["operator_id"]
            # Treat 0 as None if needed, or just pass it. Assuming 0 means "no operator" -> None
            if op_id == 0:
                op_id = None
            updates.append("operator_id = :operator_id")
            params["operator_id"] = op_id

        # Emp ID (Auto-resolve)
        emp_id_val = resolve_emp_id(current_user)

        if not emp_id_val:
            return error_resp("Unable to resolve emp_id from token", 401)

        # Tank ID (Resolve Number)
        if "tank_id" in update_data and update_data["tank_id"] is not None:
            tid = update_data["tank_id"]
            tn_row = db.execute(text("SELECT tank_number FROM tank_details WHERE tank_id = :tid LIMIT 1"), {"tid": tid}).fetchone()
            if not tn_row:
                return error_resp(f"Tank not found for id: {tid}", 404)
            
            tank_num = tn_row._mapping.get("tank_number") if hasattr(tn_row, "_mapping") else tn_row[0]
            updates.append("tank_id = :tank_id")
            updates.append("tank_number = :tank_number")
            params["tank_id"] = tid
            params["tank_number"] = tank_num

        # --- Handle Standard Fields ---
        fields_to_update = [
            "inspection_date", "status_id", "inspection_type_id", "location_id",
            "working_pressure", "frame_type", "design_temperature", "cabinet_type", "mfgr",
            "notes", "ownership", "safety_valve_brand_id", "safety_valve_model_id", "safety_valve_size_id",
            "vacuum_reading", "vacuum_uom", "lifter_weight_value", "pi_next_inspection_date"
        ]

        for field in fields_to_update:
            if field in update_data:
                val = update_data[field]
                
                # Special Logic for Safety Valve Fields: Force invalid/empty to None
                if field in ["safety_valve_brand_id", "safety_valve_model_id", "safety_valve_size_id"]:
                    if not is_valid_id(val):
                        val = None
                
                # Special Logic for other IDs: Treat 0 as None if desired (based on user request)
                if field in ["status_id", "inspection_type_id", "location_id"]:
                     if val == 0:
                         val = None

                updates.append(f"{field} = :{field}")
                params[field] = val

        # --- Execute Update & Validate ---
        if updates:
            sql = f"UPDATE tank_inspection_details SET {', '.join(updates)}, updated_at = NOW() WHERE inspection_id = :id"
            
            try:
                # Validation: Check if provided IDs exist (Only if they are NOT None)
                
                # Check Model
                if "safety_valve_model_id" in params:
                    mid = params["safety_valve_model_id"]
                    if mid is not None: # Strict None check
                        exists = db.execute(text("SELECT 1 FROM safety_valve_model WHERE id = :id LIMIT 1"), {"id": mid}).fetchone()
                        if not exists:
                            return error_resp(f"Invalid safety_valve_model_id: {mid}", 400)

                # Check Size
                if "safety_valve_size_id" in params:
                    sid = params["safety_valve_size_id"]
                    if sid is not None: # Strict None check
                        exists = db.execute(text("SELECT 1 FROM safety_valve_size WHERE id = :id LIMIT 1"), {"id": sid}).fetchone()
                        if not exists:
                            return error_resp(f"Invalid safety_valve_size_id: {sid}", 400)

                # Run Update
                db.execute(text(sql), params)
                
                # --- SYNCHRONIZE TANK MASTER ---
                # If safety_valve_brand_id was provided, update the tank_details table by tank_number
                if "safety_valve_brand_id" in params:
                    target_tank_number = params.get("tank_number", current_tank_number)
                    db.execute(
                        text("UPDATE tank_details SET safety_valve_brand_id = :svb WHERE tank_number = :tn"),
                        {"svb": params["safety_valve_brand_id"], "tn": target_tank_number}
                    )

                db.commit()
                
            except Exception as e:
                db.rollback()
                logger.error(f"DB Error during update: {e}", exc_info=True)
                raise e

        return success_resp("Inspection details updated", {"inspection_id": inspection_id}, 200)

    except Exception as e:
        logger.error(f"Error updating tank inspection details {inspection_id}: {e}", exc_info=True)
        return error_resp("Error updating inspection details", 500)


# File: /mnt/data/tank_inspection_router.py
# Replace the existing @router.get("/review/{inspection_id}") handler with this complete function.


@router.get("/review/{inspection_id}")
def get_inspection_review(
    inspection_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    try:
        # ---------------------------------------------------------
        # PERMISSION CHECK
        # ---------------------------------------------------------
        row = db.execute(
            text("SELECT is_submitted FROM tank_inspection_details WHERE inspection_id = :id"),
            {"id": inspection_id}
        ).fetchone()

        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)

        is_submitted = int(row[0])
        role_id = current_user.role_id

        if is_submitted == 1 and role_id == 2:
            return error_resp("Access denied for submitted reports", 403)

        # ---------------------------------------------------------
        # 1. FETCH INSPECTION DETAILS
        # ---------------------------------------------------------
        inspection_row = db.execute(
            text("""
                SELECT 
                    ti.inspection_id,
                    ti.inspection_date,
                    ti.report_number,
                    ti.tank_id,
                    t.tank_number,
                    ti.status_id,
                    t.mfgr,
                    t.ownership,
                    t.cabinet_type,
                    t.product_id,
                    ti.inspection_type_id,
                    ti.location_id,
                    t.safety_valve_brand_id,
                    ti.safety_valve_model_id,
                    ti.safety_valve_size_id,
                    ti.working_pressure,
                    ti.design_temperature,
                    ti.frame_type,
                    pl.location_name,
                    pit.inspection_type_name,
                    pm.product_name,
                    ps.status_name,
                    sb.brand_name AS safety_valve_brand,
                    ti.vacuum_reading,
                    ti.vacuum_uom,
                    ti.lifter_weight_value,
                    ti.pi_next_inspection_date,
                    ti.is_reviewed,
                    ti.reviewed_by,
                    ti.is_submitted,
                    ti.web_submitted
                FROM tank_inspection_details ti
                LEFT JOIN tank_details t ON ti.tank_id = t.tank_id
                LEFT JOIN location_master pl ON ti.location_id = pl.id
                LEFT JOIN inspection_type pit ON ti.inspection_type_id = pit.id
                LEFT JOIN product_master pm ON t.product_id = pm.id
                LEFT JOIN tank_status ps ON ti.status_id = ps.id
                LEFT JOIN safety_valve_brand sb ON t.safety_valve_brand_id = sb.id
                WHERE ti.inspection_id = :iid
            """),
            {"iid": inspection_id}
        ).fetchone()

        if not inspection_row:
            return error_resp("Inspection not found", 404)

        inspection = dict(inspection_row._mapping)

        # ---------------------------------------------------------
        # Format report_number dynamically to insert inspection_type
        # ---------------------------------------------------------
        report_num = inspection.get("report_number")
        itype_name = inspection.get("inspection_type_name")
        if report_num and itype_name and report_num.startswith("SG-T1-"):
            normalized_name = itype_name.upper().replace("-", "").replace(" ", "")
            if normalized_name in ["ONHIRE", "OFFHIRE", "CONDITION"]:
                inspection["report_number"] = report_num.replace("SG-T1-", f"SG-{normalized_name}-T1-", 1)


        # 1.a FETCH next_inspection_date FROM tank_certificate if not in ti
        # ---------------------------------------------------------
        if not inspection.get("pi_next_inspection_date"):
            try:
                cert_row = db.execute(
                    text("""
                        SELECT next_insp_date
                        FROM tank_certificate
                        WHERE tank_number = :tank_number
                        ORDER BY id DESC
                        LIMIT 1
                    """),
                    {"tank_number": inspection.get("tank_number")}
                ).fetchone()

                if cert_row:
                    inspection["pi_next_inspection_date"] = (
                        cert_row._mapping["next_insp_date"] if hasattr(cert_row, "_mapping") else cert_row[0]
                    )

            except Exception as e:
                logger.warning(f"Failed to fetch pi_next_inspection_date fallback: {e}")

        # ---------------------------------------------------------
        # 2. FETCH IMAGES
        # ---------------------------------------------------------
        raw_images = db.execute(
            text("""
                SELECT image_id, image_type, image_path, thumbnail_path, is_marked, is_assigned
                FROM tank_images
                WHERE inspection_id = :iid
            """),
            {"iid": inspection_id}
        ).fetchall()

        images = []
        image_map = {} # Map image_id (type ID) to URL
        for r in raw_images:
            img = dict(r._mapping)

            if img.get("image_path"):
                img["image_url"] = to_cdn_url(img["image_path"])
            if img.get("thumbnail_path"):
                img["thumbnail_url"] = to_cdn_url(img["thumbnail_path"])

            images.append(img)
            
            # Map type ID to image URL for assignment lookups
            if img.get("image_id") is not None:
                tid = str(img["image_id"])
                if tid not in image_map:
                    image_map[tid] = []
                image_map[tid].append(img.get("image_url"))

        # ---------------------------------------------------------
        # 3. FETCH CHECKLIST
        # ---------------------------------------------------------
        checklist_rows = db.execute(
            text("""
                SELECT
                    ic.id,
                    ic.job_id,
                    ic.job_name,
                    ic.sub_job_id,
                    ic.sub_job_description,
                    ic.sn,
                    ic.status_id,
                    ic.comment,
                    ic.image_id_assigned,
                    s.status_name
                FROM inspection_checklist ic
                LEFT JOIN inspection_status s ON ic.status_id = s.status_id
                WHERE ic.inspection_id = :iid
                ORDER BY ic.job_id, ic.id
            """),
            {"iid": inspection_id}
        ).fetchall()

        from collections import OrderedDict
        job_groups = OrderedDict()

        for row in checklist_rows:
            r = dict(row._mapping)
            job_id = r["job_id"]

            if job_id not in job_groups:
                job_groups[job_id] = {
                    "job_id": job_id,
                    "title": r["job_name"],
                    "items": [],
                    "status_id": 1
                }

            if r["status_id"] == 2:
                job_groups[job_id]["status_id"] = 2
            elif r["status_id"] == 3:
                job_groups[job_id]["status_id"] = 1

            job_groups[job_id]["items"].append({
                "sn": r["sn"],
                "title": r["sub_job_description"],
                "sub_job_id": r["sub_job_id"],
                "status_id": r["status_id"],
                "status_name": r["status_name"],
                "comment": r["comment"],
                "image_id_assigned": r["image_id_assigned"],
                "assigned_images": []
            })
            
            # Link image URLs if assigned
            if r.get("image_id_assigned"):
                assigned_ids = [v.strip() for v in str(r["image_id_assigned"]).split(",") if v.strip()]
                for aid in assigned_ids:
                    if aid in image_map:
                        # Find the first image of this type to get the name
                        img_name = ""
                        # Find in raw_images to get image_type
                        for raw_img in raw_images:
                            if str(raw_img._mapping.get("image_id")) == aid:
                                img_name = raw_img._mapping.get("image_type")
                                break
                        
                        for url in image_map[aid]:
                            job_groups[job_id]["items"][-1]["assigned_images"].append({
                                "url": url,
                                "name": img_name
                            })

        inspection_checklist = list(job_groups.values())

        for job in inspection_checklist:
            sid = job["status_id"]
            job["status_name"] = (
                "OK" if sid == 1 else
                "NA" if sid == 2 else
                "Flagged" if sid == 3 else
                "Unknown"
            )

        # ---------------------------------------------------------
        # 4. FINAL RESPONSE
        # ---------------------------------------------------------
        return success_resp(
            "Inspection review fetched",
            {
                "inspection": inspection,
                "images": images,
                "inspection_checklist": inspection_checklist
            },
            200
        )

    except Exception as e:
        logger.exception("Inspection review error")
        return error_resp(str(e), 500)

@router.delete("/review/{inspection_id}")
def delete_inspection_review(inspection_id: int, db: Session = Depends(get_db), current_user: Optional[dict] = Depends(get_current_user)):
    try:
        row = db.execute(text("SELECT * FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        if not row:
            return error_resp("Inspection not found", 404)
        try:
            try:
                if hasattr(row, "_mapping"):
                    insp = dict(row._mapping)
                elif isinstance(row, dict):
                    insp = row
                else:
                    insp = dict((k, v) for k, v in row)
            except Exception:
                insp = jsonable_encoder(row)

            # Delete related checklist and to-do items (will cascade delete due to FK constraints)
            try:
                db.execute(text("DELETE FROM inspection_checklist WHERE inspection_id = :iid"), {"iid": str(inspection_id)})
                db.execute(text("DELETE FROM to_do_list WHERE inspection_id = :iid"), {"iid": str(inspection_id)})
            except Exception:
                db.rollback()
            db.execute(text("DELETE FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id})
            db.commit()
            return success_resp("Inspection and related checklist/to-do entries deleted", {"inspection_id": inspection_id}, 200)
        except Exception as e:
            db.rollback()
            logger.error(f"Error deleting review for {inspection_id}: {e}", exc_info=True)
            return error_resp(str(e), 500)
    except Exception as e:
        logger.error(f"Unexpected error deleting review for {inspection_id}: {e}", exc_info=True)
        return error_resp("Internal server error", 500)



@router.delete("/delete/inspection_details/{inspection_id}")
def delete_inspection_details(inspection_id: int, db: Session = Depends(get_db), current_user: Optional[dict] = Depends(get_current_user)):
    try:
        row = db.execute(text("SELECT inspection_id FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)
        try:
            db.execute(text("DELETE FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id})
            db.commit()
            return success_resp("Inspection deleted", {"inspection_id": inspection_id}, 200)
        except Exception as e:
            db.rollback()
            logger.error(f"Error deleting inspection {inspection_id}: {e}", exc_info=True)
            return error_resp("Error deleting inspection", 500)
    except Exception as e:
        logger.error(f"Unexpected error deleting inspection {inspection_id}: {e}", exc_info=True)
        return error_resp("Internal server error", 500)


# -------------------------
# Tank details endpoint (keeps unfilled detection logic unchanged)
# -------------------------
@router.get("/tank-details/{tank_id}")
def get_tank_details(
    tank_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    try:
        # ---------------------------------------------------------
        # 1️⃣ Get tank_number from tank_details using tank_id
        # ---------------------------------------------------------
        tn_row = db.execute(
            text("""
                SELECT tank_number
                FROM tank_details
                WHERE tank_id = :tid
                LIMIT 1
            """),
            {"tid": tank_id}
        ).fetchone()

        if not tn_row:
            return error_resp(f"Tank not found for id: {tank_id}", 404)

        tank_number = (
            tn_row._mapping["tank_number"]
            if hasattr(tn_row, "_mapping")
            else tn_row[0]
        )

        if not tank_number:
            return error_resp("Tank number not found", 404)

        # ---------------------------------------------------------
        # 2️⃣ Fetch tank master details
        # ---------------------------------------------------------
        td_row = db.execute(
            text("""
                SELECT
                    t.working_pressure,
                    t.design_temperature,
                    t.frame_type,
                    t.cabinet_type,
                    t.mfgr,
                    t.ownership,
                    sb.brand_name AS safety_valve_brand_name
                FROM tank_details t
                LEFT JOIN safety_valve_brand sb ON t.safety_valve_brand_id = sb.id
                WHERE t.tank_number = :tn
                LIMIT 1
            """),
            {"tn": tank_number}
        ).fetchone()

        if not td_row:
            return error_resp(f"Tank details not found: {tank_number}", 404)

        row = (
            td_row._mapping
            if hasattr(td_row, "_mapping")
            else {
                "working_pressure": td_row[0],
                "design_temperature": td_row[1],
                "frame_type": td_row[2],
                "cabinet_type": td_row[3],
                "mfgr": td_row[4],
                "ownership": td_row[5],
                "safety_valve_brand_name": td_row[6],
            }
        )

        # ---------------------------------------------------------
        # 3️⃣ Fetch PI NEXT INSPECTION DATE
        # ---------------------------------------------------------
        pi_row = db.execute(
            text("""
                SELECT next_insp_date
                FROM tank_certificate
                WHERE tank_number = :tn
                ORDER BY id DESC
                LIMIT 1
            """),
            {"tn": tank_number}
        ).fetchone()

        pi_next_inspection_date = None
        if pi_row:
            pi_next_inspection_date = (
                pi_row._mapping["next_insp_date"]
                if hasattr(pi_row, "_mapping")
                else pi_row[0]
            )

        # ---------------------------------------------------------
        # 4️⃣ Response (NO inspection_id here)
        # ---------------------------------------------------------
        def conv(v):
            return float(v) if isinstance(v, Decimal) else v

        data = {
            "tank_id": tank_id,
            "tank_number": tank_number,
            "working_pressure": conv(row.get("working_pressure")),
            "design_temperature": conv(row.get("design_temperature")),
            "frame_type": row.get("frame_type"),
            "cabinet_type": row.get("cabinet_type"),
            "mfgr": row.get("mfgr"),
            "ownership": row.get("ownership"),
            "safety_valve_brand_name": row.get("safety_valve_brand_name"),
            "pi_next_inspection_date": pi_next_inspection_date,
        }

        return success_resp("Tank details fetched", data, 200)

    except Exception as e:
        logger.error("Error fetching tank details", exc_info=True)
        return error_resp("Error fetching tank details", 500)
@router.get("/get/inspection/{inspection_id}")
def get_inspection_by_id(
    inspection_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user),
):
    """
    Fetch inspection record by inspection_id.

    RULES:
    - Inspection must belong to logged-in emp_id
    - is_submitted = 0  → return stored values
    - is_submitted = 1  → return empty values
    """

    try:
        # ---------------------------------------------------------
        # 1️⃣ Resolve emp_id from token
        # ---------------------------------------------------------
        emp_id_val = resolve_emp_id(current_user)
        if not emp_id_val:
            return error_resp("User authentication required", 401)

        role_id = current_user.role_id

        # ---------------------------------------------------------
        # 2️⃣ Fetch inspection row
        # ---------------------------------------------------------
        row = db.execute(
            text("""
                SELECT 
                    inspection_id,
                    tank_id,
                    tank_number,
                    report_number,
                    inspection_date,
                    status_id,
                    product_id,
                    inspection_type_id,
                    location_id,
                    working_pressure,
                    design_temperature,
                    frame_type,
                    cabinet_type,
                    mfgr,
                    pi_next_inspection_date,
                    safety_valve_brand_id,
                    safety_valve_model_id,
                    safety_valve_size_id,
                    notes,
                    created_by,
                    operator_id,
                    emp_id,
                    ownership,
                    lifter_weight,
                    vacuum_reading,
                    lifter_weight_value,
                    is_submitted,
                    web_submitted
                FROM tank_inspection_details
                WHERE inspection_id = :id
                LIMIT 1
            """),
            {"id": inspection_id},
        ).fetchone()

        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)

        # ---------------------------------------------------------
        # 3️⃣ Normalize row → dict
        # ---------------------------------------------------------
        if hasattr(row, "_mapping"):
            row_dict = dict(row._mapping)
        elif isinstance(row, dict):
            row_dict = row
        else:
            row_dict = dict(row)

        # ---------------------------------------------------------
        # 4️⃣ Ownership check (emp_id) - REMOVED PER USER REQUEST
        # ---------------------------------------------------------
        # ---------------------------------------------------------
        # 5️⃣ Submission logic with role-based access
        # ---------------------------------------------------------
        is_submitted = int(row_dict.get("is_submitted", 0))
        web_submitted = int(row_dict.get("web_submitted", 0))

        if is_submitted == 1 or web_submitted == 1:
            if role_id == 2:
                return error_resp("Access denied for submitted reports", 403)
            elif role_id not in [1, 3, 4]:
                # Return empty values for editing
                return success_resp(
                    "Inspection already submitted",
                    {
                        "inspection_id": None,
                        "tank_id": row_dict.get("tank_id"),
                        "status_id": None,
                        "product_id": None,
                        "inspection_type_id": None,
                        "location_id": None,
                        "safety_valve_brand_id": None,
                        "safety_valve_model_id": None,
                        "safety_valve_size_id": None,
                        "notes": None,
                        "operator_id": None,
                        "inspection_date": None,
                        "report_number": None,
                        "lifter_weight": None,
                        "vacuum_reading": None,
                        "vacuum_uom": None,
                        "lifter_weight_value": None,
                        "is_submitted": 1,
                    },
                    200,
                )
            # For role 1,3,4, fall through to return full values

        # 🟢 NOT SUBMITTED or ALLOWED SUBMITTED → return ACTUAL STORED VALUES
        # Fetch fresh tank details (to display authoritative values in report header)
        try:
            td = fetch_tank_details(db, row_dict.get("tank_number"))
        except Exception:
            td = {}

        # Also fetch latest PI next inspection date from tank_certificate (if available)
        td_pi = fetch_pi_next_inspection_date(db, row_dict.get("tank_number"))

        # Auto-update tank_inspection_details with the latest certificate date if it differs
        current_pi = row_dict.get("pi_next_inspection_date")
        if td_pi and td_pi != current_pi:
            try:
                db.execute(
                    text("UPDATE tank_inspection_details SET pi_next_inspection_date = :td_pi, updated_at = NOW() WHERE inspection_id = :id"),
                    {"td_pi": td_pi, "id": row_dict.get("inspection_id")}
                )
                db.commit()
                row_dict["pi_next_inspection_date"] = td_pi
            except Exception as e:
                db.rollback()
                logger.error(f"Failed to auto-update pi_next_inspection_date for inspection {row_dict.get('inspection_id')}: {e}")

        def conv(v):
            return float(v) if isinstance(v, Decimal) else v

        resp = {
            "inspection_id": row_dict.get("inspection_id"),
            "tank_id": row_dict.get("tank_id"),
            "status_id": row_dict.get("status_id"),
            "product_id": row_dict.get("product_id"),
            "inspection_type_id": row_dict.get("inspection_type_id"),
            "location_id": row_dict.get("location_id"),
            "safety_valve_brand_id": row_dict.get("safety_valve_brand_id"),
            "safety_valve_model_id": row_dict.get("safety_valve_model_id"),
            "safety_valve_size_id": row_dict.get("safety_valve_size_id"),
            "notes": row_dict.get("notes"),
            "operator_id": row_dict.get("operator_id"),
            "inspection_date": row_dict.get("inspection_date"),
            "report_number": row_dict.get("report_number"),
            "lifter_weight": row_dict.get("lifter_weight"),
            "vacuum_reading": row_dict.get("vacuum_reading"),
            "vacuum_uom": row_dict.get("vacuum_uom"),
            "lifter_weight_value": row_dict.get("lifter_weight_value"),
            "is_submitted": is_submitted,
            # -- authoritative tank details (sourced from tank_details table)
            "working_pressure": conv(td.get("working_pressure") if td else row_dict.get("working_pressure")),
            "design_temperature": td.get("design_temperature") if td else row_dict.get("design_temperature"),
            "frame_type": td.get("frame_type") if td else row_dict.get("frame_type"),
            "cabinet_type": td.get("cabinet_type") if td else row_dict.get("cabinet_type"),
            "mfgr": td.get("mfgr") if td else row_dict.get("mfgr"),
            "ownership": td.get("ownership") if td else row_dict.get("ownership"),
            "pi_next_inspection_date": td_pi if td_pi is not None else row_dict.get("pi_next_inspection_date"),
        }

        return success_resp("Inspection fetched successfully", resp, 200)

    except Exception as e:
        logger.error(f"Error fetching inspection {inspection_id}: {e}", exc_info=True)
        return error_resp("Internal server error", 500)

# -------------------------
@router.get("/inspection/latest-draft/{tank_id}")
def get_latest_draft_inspection(
    tank_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user),
):
    """
    Return the latest unsubmitted inspection_id for the given tank_id
    or inspection_id = None if no draft exists.
    """
    try:
        query_sql = """
            SELECT inspection_id
            FROM tank_inspection_details
            WHERE tank_id = :tid
              AND (is_submitted = 0 OR is_submitted IS NULL)
              AND (web_submitted = 0 OR web_submitted IS NULL)
              AND (is_reviewed = 0 OR is_reviewed IS NULL)
            ORDER BY inspection_id DESC LIMIT 1
        """
        params = {"tid": tank_id}

        row = db.execute(text(query_sql), params).fetchone()

        inspection_id = None
        if row:
            if hasattr(row, "_mapping"):
                inspection_id = row._mapping.get("inspection_id")
            elif isinstance(row, dict):
                inspection_id = row.get("inspection_id")
            else:
                inspection_id = row[0]

        return success_resp(
            "Latest draft inspection fetched",
            {"inspection_id": inspection_id},
            200,
        )
    except Exception as e:
        logger.error(f"Error fetching latest draft inspection for tank {tank_id}: {e}", exc_info=True)
        return error_resp("Error fetching latest draft inspection", 500)
# -------------------------
# Delete lifter weight endpoint (keeps same semantics)
# -------------------------


# ----------------------------
# SUBMIT INSPECTION (Finalize)
# ----------------------------
@router.get("/submit")
def submit_inspection(
    inspection_id: int = Header(..., alias="inspection-id"),
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """
    Finalize the inspection.
    First validates that all required data is complete (no null values, all images present, to_do_list empty).
    Only submits if validation passes.
    Sets the status_id to 4 (Completed), is_submitted to 1, and updates the timestamp.
    """
# THEN run validation
# THEN update status_id / is_submitted

    try:
        # 1. Verify Inspection Exists
        row = db.execute(text("SELECT inspection_id FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)

        # 2. Run Validation Check (same logic as validation endpoint)
        issues = {"inspection": [], "checklist": [], "to_do_list": [], "images": []}
        
        # 2a. Check inspection row
        try:
            insp_row = db.execute(
                text("""
                    SELECT 
                        ti.inspection_id, ti.tank_id, ti.tank_number, ti.report_number, ti.inspection_date,
                        ti.status_id, ti.inspection_type_id, ti.location_id,
                        ti.vacuum_reading, ti.vacuum_uom, ti.lifter_weight_value,
                        ti.notes, ti.operator_id, ti.emp_id, ti.ownership,
                        t.product_id as product_id,
                        COALESCE(ti.pi_next_inspection_date, cert.next_insp_date) as pi_next_inspection_date
                    FROM tank_inspection_details ti
                    LEFT JOIN tank_details t ON ti.tank_id = t.tank_id
                    LEFT JOIN (
                        SELECT tank_number, MAX(next_insp_date) as next_insp_date
                        FROM tank_certificate
                        GROUP BY tank_number
                    ) cert ON ti.tank_number = cert.tank_number
                    WHERE ti.inspection_id = :id
                    LIMIT 1
                """),
                {"id": inspection_id}
            ).fetchone()

            if hasattr(insp_row, "_mapping"):
                insp = dict(insp_row._mapping)
            elif isinstance(insp_row, dict):
                insp = insp_row
            else:
                try:
                    insp = dict(zip(insp_row.keys(), insp_row))
                except Exception:
                    insp = {}

            required_inspection_fields = [
                "tank_id", "tank_number", "report_number", "inspection_date",
                "status_id", "product_id", "inspection_type_id", "location_id",
                "vacuum_reading", "lifter_weight_value",
            ]

            for f in required_inspection_fields:
                v = insp.get(f)
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    issues["inspection"].append({"field": f, "reason": "null or empty"})
                else:
                    if isinstance(v, (int, float)) and int(v) == 0:
                        issues["inspection"].append({"field": f, "reason": "zero or invalid"})

            # Smart validation: Check if next inspection date was found either in inspection or certificates
            pi_next = insp.get("pi_next_inspection_date")
            if pi_next is None or (isinstance(pi_next, str) and pi_next.strip() == ""):
                issues["inspection"].append({"field": "pi_next_inspection_date", "reason": "null or empty"})

        except Exception as e:
            logger.exception("Error validating inspection: %s", e)
            return error_resp(f"Error validating inspection: {e}", 500)

        # 2b. Validate inspection_checklist
        try:
            checklist_rows = db.execute(text("SELECT * FROM inspection_checklist WHERE inspection_id = :id"), {"id": str(inspection_id)}).fetchall() or []
            if not checklist_rows:
                issues["checklist"].append({"reason": "no checklist rows found for this inspection"})
            else:
                for r in checklist_rows:
                    rr = dict(r._mapping) if hasattr(r, "_mapping") else dict(zip(r.keys(), r))
                    row_issue = {"id": rr.get("id")}
                    for f in ("job_id", "sub_job_id", "sn", "status_id"):
                        v = rr.get(f)
                        if v is None or (isinstance(v, str) and v.strip() == ""):
                            row_issue.setdefault("missing_fields", []).append(f)
                    if "missing_fields" in row_issue:
                        issues["checklist"].append(row_issue)
        except Exception as e:
            logger.exception("Error validating checklist: %s", e)
            return error_resp(f"Error validating checklist: {e}", 500)

        # 2c. Validate to_do_list is empty
        try:
            todo_rows = db.execute(text("""
                SELECT DISTINCT c.job_id, c.job_name, t.status_id
                FROM to_do_list t
                LEFT JOIN inspection_checklist c ON t.checklist_id = c.id
                WHERE t.inspection_id = :id AND t.status_id = 2
                ORDER BY c.job_id
            """), {"id": inspection_id}).fetchall() or []
            
            if todo_rows:
                flagged_jobs = []
                for r in todo_rows:
                    rr = dict(r._mapping) if hasattr(r, "_mapping") else dict(zip(r.keys(), r))
                    job_id = rr.get("job_id")
                    job_name = rr.get("job_name")
                    if job_id is not None:
                        flagged_jobs.append({
                            "job_id": str(job_id),
                            "job_name": job_name or "",
                            "status_id": 2
                        })
                
                if flagged_jobs:
                    issues["to_do_list"] = [{
                        "reason": "to_do_list not empty - inspection has flagged items",
                        "flagged_jobs": flagged_jobs
                    }]
        except Exception as e:
            logger.exception("Error validating to_do_list: %s", e)

        # 2d. Validate images
        try:
            img_rows = db.execute(text("SELECT image_type, image_path, thumbnail_path, image_id FROM tank_images WHERE inspection_id = :id"), {"id": inspection_id}).fetchall() or []
            img_count = len(img_rows)
            
            expected_types = db.execute(text("SELECT id, image_type, count FROM image_type")).fetchall() or []
            expected_total_images = 0
            for et in expected_types:
                if hasattr(et, "_mapping"):
                    cnt = et._mapping.get("count") or 1
                elif isinstance(et, dict):
                    cnt = et.get("count") or 1
                else:
                    try:
                        _, _, cnt = et
                    except Exception:
                        cnt = 1
                expected_total_images += int(cnt)

            if expected_total_images == 0:
                expected_total_images = 15
            
            if img_count < expected_total_images:
                issues["images"].append({"reason": f"insufficient images: found {img_count}, expected {expected_total_images}"})
            else:
                for idx, r in enumerate(img_rows):
                    rr = dict(r._mapping) if hasattr(r, "_mapping") else dict(zip(r.keys(), r))
                    if not rr.get("image_path"):
                        issues["images"].append({"index": idx, "reason": "image_path missing"})
                    if (not rr.get("image_id")) and (not rr.get("image_type")):
                        issues["images"].append({"index": idx, "reason": "image type missing"})
        except Exception as e:
            logger.exception("Error validating images: %s", e)

        # 3. Check if any issues found
        any_issues = any(len(v) > 0 for v in issues.values())
        if any_issues:
            # Return detailed validation errors
            return JSONResponse(
                status_code=400,
                content={
                    "success": False,
                    "message": "Cannot submit inspection - validation failed. Please complete all required fields.",
                    "data": {
                        "inspection_id": inspection_id,
                        "issues": issues
                    }
                }
            )

        # 4. Update Status to 4 (Completed) and is_submitted = 1
        # 4. Update Status logic based on role_id
        role_id = current_user.role_id
        
        if role_id in [1, 3, 4]:
            # For Admin/Mgmt: web_submitted = 1, is_submitted remains 0
            # status_id is NOT finalized to 4 yet (because is_submitted is not 1)
             db.execute(text("""
                UPDATE tank_inspection_details 
                SET web_submitted = 1, updated_at = NOW() 
                WHERE inspection_id = :id
            """), {"id": inspection_id})
             msg = "Inspection web-submitted successfully"

        elif role_id == 2:
            # For Operator: is_submitted = 1, web_submitted = 0
            # update status to Completed (4) as well if that was the original logic for operators?
            # User said: "is_sumitted tuns to 1 AND WEB_SUBMITTED = 0"
            # And previously we had status_id = 4. Let's keep status_id updates out if user didn't ask, 
            # BUT the previous code set status_id=4. The user said specifically what to set.
            # I will follow strictly: "is_sumitted tuns to 1 AND WEB_SUBMITTED = 0"
            # The previous status_id=4 change was removed in step 561 per user request.
            
            db.execute(text("""
                UPDATE tank_inspection_details 
                SET is_submitted = 1, web_submitted = 0, updated_at = NOW() 
                WHERE inspection_id = :id
            """), {"id": inspection_id})
            msg = "Inspection submitted successfully"
            
        else:
             # Fallback for unknown roles (treat as operator or error? Defaulting to is_submitted=1 for safety if that was old behavior)
             db.execute(text("""
                UPDATE tank_inspection_details 
                SET is_submitted = 1, updated_at = NOW() 
                WHERE inspection_id = :id
            """), {"id": inspection_id})
             msg = "Inspection submitted successfully"
        
        db.commit()
        
        return success_resp(msg, {"inspection_id": inspection_id, "status": "Submitted"}, 200)

    except Exception as e:
        db.rollback()
        logger.error(f"Error submitting inspection {inspection_id}: {e}", exc_info=True)
        return error_resp("Failed to submit inspection", 500)

# ----------------------------
# REVIEW INSPECTION (Mark reviewed)
# ----------------------------
@router.post("/review_finalize/{inspection_id}")
def review_finalize_inspection(
    inspection_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """
    Mark the inspection as reviewed.
    Sets is_reviewed = 1 and reviewed_by = current_user.emp_id.
    """
    emp_id_val = resolve_emp_id(current_user)
    if not emp_id_val:
        return error_resp("Unauthorized", 401)

    try:
        # Check if inspection exists
        row = db.execute(text("SELECT inspection_id FROM tank_inspection_details WHERE inspection_id = :id"), {"id": inspection_id}).fetchone()
        if not row:
            return error_resp(f"Inspection {inspection_id} not found", 404)

        # Update is_reviewed and reviewed_by
        db.execute(
            text("UPDATE tank_inspection_details SET is_reviewed = 1, reviewed_by = :reviewer, updated_at = NOW() WHERE inspection_id = :id"),
            {"reviewer": emp_id_val, "id": inspection_id}
        )
        db.commit()

        # Insert into inspection_history
        try:
            # Fetch the updated record
            # JOIN with tank_details to get product_id and safety_valve_brand_id which aren't in ti table
            record = db.execute(
                text("""
                    SELECT ti.*, t.product_id, t.safety_valve_brand_id
                    FROM tank_inspection_details ti
                    LEFT JOIN tank_details t ON ti.tank_id = t.tank_id
                    WHERE ti.inspection_id = :id
                """),
                {"id": inspection_id}
            ).fetchone()

            if record:
                # Use mapping/getattr to be safe with row object
                # Some environments might return row, others mapping
                r = record._mapping if hasattr(record, "_mapping") else dict(zip(record.keys(), record))

                history_entry = InspectionHistory(
                    inspection_id=r["inspection_id"],
                    inspection_date=r["inspection_date"],
                    created_at=r["created_at"],
                    updated_at=r["updated_at"],
                    report_number=r["report_number"],
                    tank_id=r["tank_id"],
                    tank_number=r["tank_number"],
                    status_id=r["status_id"],
                    product_id=r.get("product_id"),
                    inspection_type_id=r["inspection_type_id"],
                    location_id=r["location_id"],
                    working_pressure=r["working_pressure"],
                    design_temperature=r["design_temperature"],
                    frame_type=r["frame_type"],
                    cabinet_type=r["cabinet_type"],
                    mfgr=r["mfgr"],
                    safety_valve_brand_id=r.get("safety_valve_brand_id"),
                    safety_valve_model_id=r["safety_valve_model_id"],
                    safety_valve_size_id=r["safety_valve_size_id"],
                    pi_next_inspection_date=r["pi_next_inspection_date"],
                    notes=r["notes"],
                    lifter_weight=r.get("lifter_weight"),
                    lifter_weight_thumbnail=r.get("lifter_weight_thumbnail"),
                    vacuum_reading=r["vacuum_reading"],
                    vacuum_uom=r["vacuum_uom"],
                    lifter_weight_value=r["lifter_weight_value"],
                    emp_id=r["emp_id"],
                    operator_id=r["operator_id"],
                    ownership=r["ownership"],
                    is_submitted=r["is_submitted"],
                    is_reviewed=r["is_reviewed"],
                    reviewed_by=r["reviewed_by"],
                    web_submitted=r["web_submitted"],
                    created_by=r["created_by"],
                    updated_by=r["updated_by"],
                    history_date=func.now()
                )
                db.add(history_entry)
                db.commit()
        except Exception as e:
            logger.error(f"Error inserting into inspection_history for {inspection_id}: {e}")
            # Don't fail the whole operation

        return success_resp("Inspection report marked as REVIEWED", {"inspection_id": inspection_id, "reviewed_by": emp_id_val}, 200)

    except Exception as e:
        db.rollback()
        logger.error(f"Error reviewing inspection {inspection_id}: {e}", exc_info=True)
        return error_resp(str(e), 500)

# ----------------------------
# GET INSPECTION HISTORY
# ----------------------------
@router.get("/history")
def get_inspection_history(
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    """
    Get all inspection history records.
    """
    try:
        role_id = current_user.role_id
        if role_id == 2:
            return success_resp("History", [], 200)

        results = db.execute(text("""
            SELECT 
                ih.id,
                ih.inspection_id,
                ih.inspection_date,
                ih.report_number,
                ih.tank_number,
                ih.history_date,
                ih.is_reviewed,
                ih.reviewed_by,
                ps.status_name,
                pit.inspection_type_name,
                pm.product_name,
                pl.location_name
            FROM inspection_history ih
            LEFT JOIN tank_status ps ON ih.status_id = ps.id
            LEFT JOIN inspection_type pit ON ih.inspection_type_id = pit.id
            LEFT JOIN product_master pm ON ih.product_id = pm.id
            LEFT JOIN location_master pl ON ih.location_id = pl.id
            ORDER BY ih.history_date DESC
        """)).fetchall()

        history = []
        for r in results:
            d = dict(r._mapping)
            # Format report_number dynamically to insert inspection_type
            report_num = d.get("report_number")
            itype_name = d.get("inspection_type_name")
            if report_num and itype_name and report_num.startswith("SG-T1-"):
                normalized_name = itype_name.upper().replace("-", "").replace(" ", "")
                if normalized_name in ["ONHIRE", "OFFHIRE", "CONDITION"]:
                    d["report_number"] = report_num.replace("SG-T1-", f"SG-{normalized_name}-T1-", 1)
            history.append(d)

        return success_resp("Inspection history fetched", history, 200)

    except Exception as e:
        logger.exception("Error fetching inspection history")
        return error_resp("Failed to fetch history", 500)


@router.get("/user/me")
def get_current_user_info(current_user: Optional[dict] = Depends(get_current_user)):
    """Return basic info about the currently authenticated user.

    Accepts either a `User` model instance or a decoded payload dict.
    """
    if not current_user:
        return error_resp("User not authenticated", 401)

    if hasattr(current_user, "emp_id"):
        emp_id = getattr(current_user, "emp_id", None)
        role_id = getattr(current_user, "role_id", None)
        login_name = getattr(current_user, "login_name", None)
        email = getattr(current_user, "email", None)
    else:
        emp_id = current_user.get("emp_id")
        role_id = current_user.get("role_id")
        login_name = current_user.get("login_name")
        email = current_user.get("email")

    return success_resp("User info", {
        "emp_id": emp_id,
        "role_id": role_id,
        "login_name": login_name,
        "email": email,
    })

@router.post("/copy/{inspection_id}")
def copy_inspection(
    inspection_id: int,
    new_type_id: int,
    db: Session = Depends(get_db),
    current_user: Optional[dict] = Depends(get_current_user)
):
    try:
        from app.services.copy_inspection_service import CopyInspectionService
        
        new_id, new_report_no = CopyInspectionService.copy_inspection(
            db, inspection_id, new_type_id, current_user
        )
        
        return success_resp("Inspection copied successfully", {
            "new_inspection_id": new_id,
            "new_report_number": new_report_no
        }, 201)

    except ValueError as e:
        return error_resp(str(e), 404)
    except Exception as e:
        logger.error(f"Error copying inspection {inspection_id}: {e}", exc_info=True)
        return error_resp("Failed to copy inspection", 500)


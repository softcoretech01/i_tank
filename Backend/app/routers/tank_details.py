from fastapi import APIRouter, Depends, HTTPException, Header, File, UploadFile
from app.utils.s3_utils import to_cdn_url
import os
import shutil
import uuid
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.models.users_model import User
from app.database import get_db
from app.models.tank_header import Tank
from app.models.tank_details import TankDetails
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from datetime import datetime

router = APIRouter()
    # Look up master values by ID
from app.models.manufacturer_master_model import ManufacturerMaster
from app.models.standard_master_model import StandardMaster
from app.models.tankcode_iso_master_model import TankCodeISOMaster
from app.models.ownership_master_model import OwnershipMaster
from app.models.design_temperature_master_model import DesignTemperatureMaster
from app.models.cabinet_type_master_model import CabinetTypeMaster
from app.models.frame_type_master_model import FrameTypeMaster
from app.models.size_master_model import SizeMaster
from app.models.pump_master_model import PumpMaster
from app.models.mawp_master_model import MawpMaster
from app.models.un_code_master_model import UNISOCODEMaster
from app.routers.tank_inspection_router import get_current_user
from app.models.multiple_regulation import MultipleRegulation
from app.models.regulations_master import RegulationsMaster

# --- Helper function to convert "" to None for numeric types ---
def to_float_or_none(value: any) -> float | None:
    if value == "" or value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        raise HTTPException(status_code=400, detail=f"Invalid numeric value: {value}")


def resolve_names_from_ids(db: Session, model, id_list, id_field='id', name_field='standard_name'):
    """Given a list of ids, return a list of names from the given SQLAlchemy model.
    Preserves the order of id_list and silently skips ids not found.
    """
    if not id_list:
        return []
    # normalize ids to ints
    try:
        ids = [int(x) for x in id_list]
    except Exception:
        raise HTTPException(status_code=400, detail=f"Invalid id list: {id_list}")

    rows = db.query(model).filter(getattr(model, id_field).in_(ids)).all()
    lookup = {getattr(r, id_field): getattr(r, name_field) for r in rows}
    names = [lookup.get(i) for i in ids if lookup.get(i)]
    return names

@router.post("/")
def create_tank(data: dict, db: Session = Depends(get_db),current_user: User = Depends(get_current_user),):
    # Note: frontend may send either a single `standard_id` or multiple `standard` ids (list).
    # Require at least one of them to be present. Keep other required ids as before.
    # Allow clients to provide either a size_id (master id) or a free-form `size` string.
    required_fields = [
        "tank_number", "manufacturer_id",
        "capacity_l", "mawp_id", "date_mfg", "un_code", "initial_test", "design_temperature_id", "tare_weight_kg",
        "mgw_kg", "pump_id",
        "tank_iso_code_id", "cabinet_id", "frame_type_id", "ownership_id"
    ]
    try:
        from app.models.users_model import User
    except Exception:
        User = None

    manufacturer = db.query(ManufacturerMaster).filter_by(manufacturer_id=data.get("manufacturer_id")).first()
    if not manufacturer: raise HTTPException(status_code=400, detail=f"Invalid manufacturer_id")
    
    # STANDARD: frontend may send either a single id (standard_id),
    # or multiple ids as list in `standard` (e.g. standard: [1,2]). We will resolve names.
    standard_names = []
    if data.get("standard_id"):
        s = db.query(StandardMaster).filter_by(id=data.get("standard_id")).first()
        if not s:
            raise HTTPException(status_code=400, detail=f"Invalid standard_id")
        standard_names = [s.standard_name]
    elif data.get("standard") and isinstance(data.get("standard"), list):
        standard_names = resolve_names_from_ids(db, StandardMaster, data.get("standard"), id_field='id', name_field='standard_name')
        if not standard_names:
            raise HTTPException(status_code=400, detail=f"No valid standards found for provided ids")
    elif data.get("standard") and isinstance(data.get("standard"), str):
        # allow comma-separated names from older clients
        standard_names = [s.strip() for s in data.get("standard").split(',') if s.strip()]
    else:
        # leave empty; it's optional depending on your UI
        standard_names = []
    
    tank_iso_code = db.query(TankCodeISOMaster).filter_by(id=data.get("tank_iso_code_id")).first()
    if not tank_iso_code: raise HTTPException(status_code=400, detail=f"Invalid tank_iso_code_id")
    
    # UN CODES: frontend will send selected un_code ids in `un_code` as list or as comma-separated string of codes.
    un_code_names = []
    if data.get("un_code") and isinstance(data.get("un_code"), list):
        # treat as list of ids
        un_code_names = resolve_names_from_ids(db, UNISOCODEMaster, data.get("un_code"), id_field='id', name_field='un_code')
        if not un_code_names:
            raise HTTPException(status_code=400, detail="No valid un_code ids provided")
    elif data.get("un_code") and isinstance(data.get("un_code"), str):
        # could be comma-separated codes already
        # accept both formats: comma-separated ids or codes
        s = data.get("un_code").strip()
        if s == "":
            un_code_names = []
        elif s.replace(',', '').isdigit():
            # if string of numbers separated by commas, try to resolve as ids
            ids = [p.strip() for p in s.split(',') if p.strip()]
            un_code_names = resolve_names_from_ids(db, UNISOCODEMaster, ids, id_field='id', name_field='un_iso_code')
            if not un_code_names:
                # fallback: treat original string as codes
                un_code_names = [p.strip() for p in s.split(',') if p.strip()]
        else:
            un_code_names = [p.strip() for p in s.split(',') if p.strip()]
    else:
        # if not provided, treat as empty list
        un_code_names = []
    
    ownership = db.query(OwnershipMaster).filter_by(id=data.get("ownership_id")).first()
    if not ownership: raise HTTPException(status_code=400, detail=f"Invalid ownership_id")
    
    design_temperature = db.query(DesignTemperatureMaster).filter_by(id=data.get("design_temperature_id")).first()
    if not design_temperature: raise HTTPException(status_code=400, detail=f"Invalid design_temperature_id")
    
    cabinet = db.query(CabinetTypeMaster).filter_by(id=data.get("cabinet_id")).first()
    if not cabinet: raise HTTPException(status_code=400, detail=f"Invalid cabinet_id")
    
    frame_type = db.query(FrameTypeMaster).filter_by(id=data.get("frame_type_id")).first()
    if not frame_type: raise HTTPException(status_code=400, detail=f"Invalid frame_type_id")
    
    # SIZE: accept either `size_id` (existing behavior) or a free-form `size` string.
    size_label = None
    if data.get("size_id"):
        size_row = db.query(SizeMaster).filter_by(id=data.get("size_id")).first()
        if not size_row:
            raise HTTPException(status_code=400, detail=f"Invalid size_id")
        size_label = size_row.size_label
    elif data.get("size") and isinstance(data.get("size"), str) and data.get("size").strip():
        size_label = data.get("size").strip()
    else:
        # If neither provided, it's an error because UI requires size
        raise HTTPException(status_code=400, detail=f"Either size_id or size (string) must be provided")
    
    pump = db.query(PumpMaster).filter_by(id=data.get("pump_id")).first()
    if not pump: raise HTTPException(status_code=400, detail=f"Invalid pump_id")
    
    mawp = db.query(MawpMaster).filter_by(id=data.get("mawp_id")).first()
    if not mawp: raise HTTPException(status_code=400, detail=f"Invalid mawp_id")

    # Get emp_id from the logged-in user (from JWT token)
    if not current_user or not getattr(current_user, "emp_id", None):
        raise HTTPException(
            status_code=401,
            detail="Could not resolve emp_id from logged-in user"
        )

    emp_id = str(current_user.emp_id)


    for field in required_fields:
        if field not in data or data[field] == "":
            raise HTTPException(status_code=400, detail=f"Required field '{field}' is missing or empty")

    # Standard is now optional, can be null

    existing_tank = db.query(Tank).filter(Tank.tank_number == data["tank_number"]).first()
    if existing_tank:
        raise HTTPException(status_code=400, detail=f"Tank number '{data['tank_number']}' already exists")

    tank = Tank(
        tank_number=data["tank_number"],
        created_by=emp_id,
        updated_by=emp_id
    )
    db.add(tank)
    db.commit()
    db.refresh(tank)

    status = data.get("status", "active")
    if status not in ["active", "inactive"]:
        raise HTTPException(status_code=400, detail="status must be 'active' or 'inactive'")
    tank.status = status
    db.commit()
    db.refresh(tank)
    
    try:
        # Validate that tare_weight is less than mgw
        mgw_val = to_float_or_none(data.get("mgw_kg"))
        tare_val = to_float_or_none(data.get("tare_weight_kg"))
        if mgw_val is None or tare_val is None:
            raise HTTPException(
                status_code=400,
                detail="Both 'mgw_kg' and 'tare_weight_kg' must be provided and numeric",
            )

        if not (tare_val < mgw_val):
            # Custom validation message that should be sent back to the frontend
            raise HTTPException(
                status_code=400,
                detail="tare_weight must be less than mgw_kg",
            )

        mpl_kg = mgw_val - tare_val
        tank_detail = TankDetails(
            tank_id=tank.id,
            tank_number=data["tank_number"],
            status=status,
            mfgr=manufacturer.manufacturer_name,
            date_mfg=data.get("date_mfg") or None,
            initial_test=data.get("initial_test") or None,
                # store comma-separated names (not ids)
                un_code=','.join(un_code_names) if un_code_names else (data.get("un_code") or None),
            tank_iso_code=tank_iso_code.tankcode_iso,
                standard=','.join(standard_names) if standard_names else (data.get("standard") or None),
            capacity_l=to_float_or_none(data["capacity_l"]),
            mawp=mawp.mawp_value,
            design_temperature=design_temperature.design_temperature,
            tare_weight_kg=to_float_or_none(data["tare_weight_kg"]),
            mgw_kg=to_float_or_none(data["mgw_kg"]),
            mpl_kg=mpl_kg,
            size=size_label,
            pump_type=pump.pump_value,
            gross_kg=to_float_or_none(data["mgw_kg"]),
            net_kg=mpl_kg,
            working_pressure=mawp.mawp_value,
            cabinet_type=cabinet.cabinet_type,
            frame_type=frame_type.frame_type,
            remark=data.get("remark") or None,
            ownership=ownership.ownership_name,
            created_by=emp_id,
            updated_by=emp_id,
            color_body_frame=data.get("color_body_frame"),
            evacuation_valve=data.get("evacuation_valve"),
            product_id=data.get("product_id"),
            safety_valve_brand_id=data.get("safety_valve_brand_id"),
            tank_number_image_path=data.get("tank_number_image_path")
        )

        db.add(tank_detail)
        db.commit()
        db.refresh(tank_detail)

        # Store multiple regulations
        if "regulations" in data and isinstance(data["regulations"], list):
            for reg_id in data["regulations"]:
                reg_master = db.query(RegulationsMaster).filter(RegulationsMaster.id == reg_id).first()
                if reg_master:
                    new_reg = MultipleRegulation(
                        tank_id=tank.id,
                        regulation_id=reg_id,
                        regulation_name=reg_master.regulation_name
                    )
                    db.add(new_reg)
            db.commit()

    # Let validation errors (HTTPException) pass through with their own status & message
    except HTTPException as e:
        db.rollback()
        db.delete(tank)
        db.commit()
        raise e

    # Any other unexpected error becomes a 500
    except Exception as e:
        db.rollback()
        db.delete(tank)
        db.commit()
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create tank details: {str(e)}",
        )

    return {
        "message": "Tank created successfully",
        "tank_id": tank.id,
        "data": {
            "tank_header": tank,
            "tank_details": tank_detail
        }
    }


# --- GET ALL TANKS ---

@router.post("/upload-image")
def upload_tank_image(file: UploadFile = File(...)):
    try:
        from app.utils.s3_utils import build_s3_key, upload_fileobj_to_s3
        
        file_ext = os.path.splitext(file.filename)[1]
        s3_key = build_s3_key(file.filename)
        
        upload_fileobj_to_s3(file.file, s3_key, file.content_type)
            
        return {"path": s3_key} 
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to upload image to S3: {str(e)}")


@router.get("/")
def get_all_tanks(db: Session = Depends(get_db)):
    results = db.query(Tank, TankDetails).join(TankDetails, Tank.id == TankDetails.tank_id).all()
    
    response_data = []
    for r in results:
        tank_id = r[0].id
        regulations = db.query(MultipleRegulation).filter(MultipleRegulation.tank_id == tank_id).all()
        regulation_names = ", ".join([reg.regulation_name for reg in regulations])

        response_data.append({
            "id": r[0].id,
            "tank_number": r[0].tank_number,
            "status": r[1].status,
            "mfgr": r[1].mfgr,
            "date_mfg": r[1].date_mfg,
            "initial_test": r[1].initial_test,
            "un_code": r[1].un_code, 
            "tank_iso_code": r[1].tank_iso_code,
            "standard": r[1].standard,
            "capacity_l": r[1].capacity_l,
            "mawp": r[1].mawp,
            "design_temperature": r[1].design_temperature,
            "tare_weight_kg": r[1].tare_weight_kg,
            "mgw_kg": r[1].mgw_kg,
            "mpl_kg": r[1].mpl_kg,
            "size": r[1].size,
            "pump_type": r[1].pump_type,
            "gross_kg": r[1].gross_kg,
            "net_kg": r[1].net_kg,
            "working_pressure": r[1].working_pressure,
            "cabinet_type": r[1].cabinet_type,
            "frame_type": r[1].frame_type,
            "remark": r[1].remark,
            "owner": r[1].ownership, 
            "created_by": r[0].created_by,
            "color_body_frame": r[1].color_body_frame,
            "evacuation_valve": r[1].evacuation_valve,
            "product_id": r[1].product_id,
            "safety_valve_brand_id": r[1].safety_valve_brand_id,
            "tank_number_image_path": to_cdn_url(r[1].tank_number_image_path) if r[1].tank_number_image_path else None,
            "regulations": regulation_names
        })
    
    return response_data

# --- EXPORT TO EXCEL ---
@router.get("/export-to-excel")
def export_to_excel(db: Session = Depends(get_db)):
    results = db.query(Tank, TankDetails).join(TankDetails, Tank.id == TankDetails.tank_id).all()
    
    if not results:
        raise HTTPException(status_code=404, detail="No tank details found to export")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tank Details"
    
    # Removed 'Standard' column from export per frontend requirement (allow null and hide in UI)
    headers = [
        "ID", "Tank ID", "Tank Number", "Status", "Manufacturer (MFGR)",
        "Date of Manufacture", "Initial Test", "UN Code", "Capacity (L)", "MAWP",
        "Design Temperature", "Tare Weight (kg)", "MGW (kg)", "MPL (kg)", "Size",
        "Pump Type", "Gross (kg)", "Net (kg)", "Remark", "Owner", "Created By", "Updated By"
    ]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    for row_num, (tank, tank_detail) in enumerate(results, 2):
        ws.cell(row=row_num, column=1, value=tank_detail.id)
        ws.cell(row=row_num, column=2, value=tank_detail.tank_id)
        ws.cell(row=row_num, column=3, value=tank_detail.tank_number or tank.tank_number)
        ws.cell(row=row_num, column=4, value=tank_detail.status)
        ws.cell(row=row_num, column=5, value=tank_detail.mfgr)
        ws.cell(
            row=row_num,
            column=6,
            value=tank_detail.date_mfg if tank_detail.date_mfg else None
        )
        ws.cell(row=row_num, column=7, value=tank_detail.initial_test if tank_detail.initial_test else None)
        ws.cell(row=row_num, column=8, value=tank_detail.un_code)
        # omitted: standard (user requested hidden / nullable)
        ws.cell(row=row_num, column=9, value=tank_detail.capacity_l)
        ws.cell(row=row_num, column=10, value=tank_detail.mawp)
        ws.cell(row=row_num, column=11, value=tank_detail.design_temperature)
        ws.cell(row=row_num, column=12, value=tank_detail.tare_weight_kg)
        ws.cell(row=row_num, column=13, value=tank_detail.mgw_kg)
        ws.cell(row=row_num, column=14, value=tank_detail.mpl_kg)
        ws.cell(row=row_num, column=15, value=tank_detail.size)
        ws.cell(row=row_num, column=16, value=tank_detail.pump_type)
        ws.cell(row=row_num, column=17, value=tank_detail.gross_kg)
        ws.cell(row=row_num, column=18, value=tank_detail.net_kg)
        ws.cell(row=row_num, column=19, value=tank_detail.remark)
        ws.cell(row=row_num, column=20, value=tank_detail.ownership)
        ws.cell(row=row_num, column=21, value=tank_detail.created_by)
        ws.cell(row=row_num, column=22, value=tank_detail.updated_by)
    
    for col_num, header in enumerate(headers, 1):
        column_letter = ws.cell(row=1, column=col_num).column_letter
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
            if row[0].value:
                max_length = max(max_length, len(str(row[0].value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    ws.row_dimensions[1].height = 25
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"tank_details_export_{timestamp}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# --- UPDATE TANK ---
@router.put("/{tank_id}")
def update_tank(tank_id: int, data: dict, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    tank = db.query(Tank).filter(Tank.id == tank_id).first()
    tank_detail = db.query(TankDetails).filter(TankDetails.tank_id == tank_id).first()

    if not tank or not tank_detail:
        raise HTTPException(status_code=404, detail="Tank not found")

    # Get emp_id from the logged-in user
    if not current_user or not getattr(current_user, "emp_id", None):
        raise HTTPException(
            status_code=401,
            detail="Could not resolve emp_id from logged-in user"
        )

    emp_id = str(current_user.emp_id)

    # 1. Handle Renaming
    if "tank_number" in data:
        new_tank_number = data["tank_number"]
        if new_tank_number != tank.tank_number:
            existing = db.query(Tank).filter(Tank.tank_number == new_tank_number).first()
            if existing:
                raise HTTPException(status_code=400, detail=f"Tank number '{new_tank_number}' already exists")
            
            tank_detail.tank_number = None
            tank.tank_number = new_tank_number
            db.commit()
            tank_detail.tank_number = new_tank_number
    
    tank.updated_by = emp_id

    if "status" in data:
        status = data["status"]
        if status not in ["active", "inactive"]:
            raise HTTPException(status_code=400, detail="status must be 'active' or 'inactive'")
        tank.status = status

    if "owner" in data:
        data["ownership"] = data["owner"]

    detail_fields = [
        "status", "mfgr", "date_mfg", "initial_test", "un_code",
        "capacity_l", "mawp", "design_temperature", "tare_weight_kg",
        "mgw_kg", "mpl_kg", "size", "pump_type",
        "gross_kg", "net_kg",
        "remark", "ownership",
        "working_pressure", "cabinet_type", "frame_type",
        "tank_iso_code", "standard",
        "color_body_frame", "evacuation_valve", "tank_number_image_path",
        "product_id", "safety_valve_brand_id"
    ]

    for field in detail_fields:
        if field in data:
            value = data[field]
            
            # special handling: numeric fields
            if field in ["capacity_l", "tare_weight_kg", "mgw_kg", "mpl_kg", "gross_kg", "net_kg"]:
                setattr(tank_detail, field, to_float_or_none(value))
            elif field == "mawp" or field == "working_pressure":
                # MAWP and working pressure are string fields (e.g., '21.4 bar')
                setattr(tank_detail, field, value)
            elif field == "date_mfg":
                setattr(tank_detail, field, value or None)
            elif field == "standard":
                # Accept either list of ids or comma-separated names
                if isinstance(value, list):
                    names = resolve_names_from_ids(db, StandardMaster, value, id_field='id', name_field='standard_name')
                    setattr(tank_detail, field, ','.join(names) if names else None)
                elif isinstance(value, str) and value.strip():
                    # allow comma-separated names already
                    setattr(tank_detail, field, value)
                else:
                    setattr(tank_detail, field, None)
            elif field == "un_code":
                # Accept list of ids or comma-separated codes/names
                if isinstance(value, list):
                    names = resolve_names_from_ids(db, UNISOCODEMaster, value, id_field='id', name_field='un_code')
                    setattr(tank_detail, field, ','.join(names) if names else None)
                elif isinstance(value, str) and value.strip():
                    # if csv of ids - try to resolve
                    s = value.strip()
                    if s.replace(',', '').isdigit():
                        ids = [p.strip() for p in s.split(',') if p.strip()]
                        names = resolve_names_from_ids(db, UNISOCODEMaster, ids, id_field='id', name_field='un_iso_code')
                        setattr(tank_detail, field, ','.join(names) if names else s)
                    else:
                        setattr(tank_detail, field, s)
            else:
                setattr(tank_detail, field, value)

    # Handle Multiple Regulations Sync
    if "regulations" in data and isinstance(data["regulations"], list):
        # 1. Delete old links
        db.query(MultipleRegulation).filter(MultipleRegulation.tank_id == tank_id).delete()
        
        # 2. Add new links
        for reg_id in data["regulations"]:
            reg_master = db.query(RegulationsMaster).filter(RegulationsMaster.id == reg_id).first()
            if reg_master:
                new_reg = MultipleRegulation(
                    tank_id=tank.id,
                    regulation_id=reg_id,
                    regulation_name=reg_master.regulation_name
                )
                db.add(new_reg)
        db.commit()

    # Set updated_by for tank_detail
    tank_detail.updated_by = emp_id

    # Recalculate mpl_kg, net_kg, gross_kg, working_pressure based on tare_weight_kg and mgw_kg, mawp
    tare_val = to_float_or_none(getattr(tank_detail, 'tare_weight_kg', None))
    mgw_val = to_float_or_none(getattr(tank_detail, 'mgw_kg', None))
    mawp_val = getattr(tank_detail, 'mawp', None)
    if tare_val is not None and mgw_val is not None:
        tank_detail.mpl_kg = mgw_val - tare_val
        tank_detail.net_kg = mgw_val - tare_val
        tank_detail.gross_kg = mgw_val
        if mawp_val:
            tank_detail.working_pressure = mawp_val

    # Validate that tare_weight is less than mgw after applying updates
    try:
        tare_val = to_float_or_none(getattr(tank_detail, 'tare_weight_kg', None))
        mgw_val = to_float_or_none(getattr(tank_detail, 'mgw_kg', None))
        if tare_val is not None and mgw_val is not None and not (tare_val < mgw_val):
            raise HTTPException(status_code=400, detail="tare_weight must be less than mgw")
    except HTTPException:
        # Re-raise validation errors
        raise

    try:
        db.commit()
        db.refresh(tank)
        db.refresh(tank_detail)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update tank details: {str(e)}")
    
    return {"message": "Tank updated successfully"}


@router.delete("/{tank_id}")
def delete_tank(tank_id: int, db: Session = Depends(get_db), authorization: str = Header(...)):
    tank_detail = db.query(TankDetails).filter(TankDetails.tank_id == tank_id).first()
    tank = db.query(Tank).filter(Tank.id == tank_id).first()

    if not tank:
        raise HTTPException(status_code=404, detail="Tank not found")

    if tank_detail:
        db.delete(tank_detail)
    db.delete(tank)
    db.commit()

    return {"message": "Tank deleted successfully"}

# --- GET BY ID ---
@router.get("/{tank_id}")
def get_tank_by_id(tank_id: int, db: Session = Depends(get_db)):
    tank = db.query(Tank).filter(Tank.id == tank_id).first()
    tank_detail = db.query(TankDetails).filter(TankDetails.tank_id == tank_id).first()

    if not tank or not tank_detail:
        raise HTTPException(status_code=404, detail="Tank not found")

    # Fetch Multiple Regulations
    linked_regulations = db.query(MultipleRegulation).filter(MultipleRegulation.tank_id == tank_id).all()
    regulation_ids = [reg.regulation_id for reg in linked_regulations]

    return {
        "id": tank.id,
        "tank_number": tank.tank_number,
        "initial_test": tank_detail.initial_test,
        "status": tank_detail.status,
        "mfgr": tank_detail.mfgr,
        "date_mfg": tank_detail.date_mfg,
        "un_code": tank_detail.un_code, # CHANGED: un_iso_code -> un_code
        "tank_iso_code": tank_detail.tank_iso_code,
        "standard": tank_detail.standard,
        "capacity_l": tank_detail.capacity_l,
        "mawp": tank_detail.mawp,
        "design_temperature": tank_detail.design_temperature,
        "tare_weight_kg": tank_detail.tare_weight_kg,
        "mgw_kg": tank_detail.mgw_kg,
        "mpl_kg": tank_detail.mpl_kg,
        "size": tank_detail.size,
        "pump_type": tank_detail.pump_type,
        "gross_kg": tank_detail.gross_kg,
        "net_kg": tank_detail.net_kg,
        "remark": tank_detail.remark,
        "owner": tank_detail.ownership, 
        "created_by": tank.created_by,
        "updated_by": tank.updated_by,
        "working_pressure": tank_detail.working_pressure,
        "cabinet_type": tank_detail.cabinet_type,
        "frame_type": tank_detail.frame_type,
        "color_body_frame": tank_detail.color_body_frame,
        "evacuation_valve": tank_detail.evacuation_valve,
        "product_id": tank_detail.product_id,
        "safety_valve_brand_id": tank_detail.safety_valve_brand_id,
        "tank_number_image_path": to_cdn_url(tank_detail.tank_number_image_path) if tank_detail.tank_number_image_path else None,
        "regulations": regulation_ids
    }

@router.get("/by-number/{tank_number}")
def get_tank_by_number(tank_number: str, db: Session = Depends(get_db)):
    tank = db.query(Tank).filter(Tank.tank_number == tank_number).first()
    if not tank:
        raise HTTPException(status_code=404, detail="Tank not found")

    tank_detail = db.query(TankDetails).filter(TankDetails.tank_id == tank.id).first()
    if not tank_detail:
        raise HTTPException(status_code=404, detail="Tank details not found")

    return {
        "tank_number": tank.tank_number,
        "date_mfg": tank_detail.date_mfg
    }
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Header
from sqlalchemy.orm import Session
from sqlalchemy.exc import OperationalError
from sqlalchemy import text
from app.database import get_db
from app.models.tank_certificate import TankCertificate
from app.models.tank_header import Tank
import os
from typing import Optional, Union
from datetime import date as date_type, datetime
import re
import traceback
import logging

from app.utils.upload_utils import save_uploaded_file, delete_file_if_exists
from app.utils.s3_utils import to_cdn_url
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from fastapi.responses import StreamingResponse
import jwt

JWT_SECRET = os.getenv("JWT_SECRET", "change_this_in_production")
JWT_ALGORITHM = os.getenv("JWT_ALGORITHM", "HS256")

try:
    from app.models.users_model import User
except Exception:
    User = None


def get_emp_id_from_token(authorization: Optional[str]) -> str:
    """Soft auth: parse emp_id from Bearer token. Returns 'System' on any failure."""
    if not authorization:
        return "System"
    try:
        auth = authorization.strip()
        token = auth[7:].strip() if auth.lower().startswith("bearer") else auth
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        emp_id = payload.get("emp_id")
        return str(int(emp_id)) if emp_id else "System"
    except Exception:
        return "System"

router = APIRouter()
logger = logging.getLogger(__name__)

UPLOAD_ROOT = os.getenv(
    "UPLOAD_ROOT",
    os.path.join(os.path.dirname(__file__), ".", ".", "uploads"),
)

CERTIFICATE_TYPE = "certificates"

# Max certificate size: 2 MB
MAX_certificate_SIZE = 2 * 1024 * 1024
ALLOWED_certificate_TYPES = {"application/pdf"}


# --- HELPERS ---

def validate_certificate(upload_file: UploadFile, field_label: str):
    """Validate that file is PDF and ≤ 2 MB."""
    content_type = getattr(upload_file, "content_type", "") or ""
    # Be lenient: allow octet-stream too (some browsers send this for PDFs)
    if content_type.lower() not in ALLOWED_certificate_TYPES and "pdf" not in content_type.lower():
        raise HTTPException(
            status_code=400,
            detail=f"{field_label}: Only PDF files are allowed (got '{content_type}')."
        )
    contents = upload_file.file.read()
    if len(contents) > MAX_certificate_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"{field_label}: File size exceeds 2 MB ({len(contents) // 1024} KB)."
        )
    upload_file.file.seek(0)


def clean_form_data(value: Optional[str]):
    return value.strip() if value else None


def safe_serialize_date(date_value: Union[date_type, datetime, str, None]) -> Optional[str]:
    if isinstance(date_value, str):
        return date_value
    if date_value and isinstance(date_value, (date_type, datetime)):
        return date_value.strftime("%Y/%m")
    return None


def _normalize_date_str(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    s = s.strip().replace('-', '/')
    m = re.match(r'^(\d{4})/(\d{2})/(\d{2})$', s)
    if m:
        return f"{m.group(1)}/{m.group(2)}"
    m2 = re.match(r'^(\d{4})/(\d{2})$', s)
    if m2:
        mm = int(m2.group(2))
        if 1 <= mm <= 12:
            return f"{m2.group(1)}/{m2.group(2)}"
    return None


def calculate_next_inspection_date(insp_date_str: str) -> Optional[str]:
    """Add 30 months (2.5 years) to a YYYY/MM or YYYY-MM string."""
    norm = _normalize_date_str(insp_date_str)
    if not norm:
        return None
    try:
        # Normalize returns YYYY/MM
        year, month = map(int, norm.split('/'))
        # Add 30 months
        total_months = year * 12 + (month - 1) + 30
        new_year = total_months // 12
        new_month = (total_months % 12) + 1
        return f"{new_year}/{new_month:02d}"
    except Exception:
        return None


def to_url(raw):
    if not raw:
        return ""
    return to_cdn_url(raw) if "://" not in raw else raw


def serialize_certificate(cert):
    # Support both SQLAlchemy Row objects (dot notation) and RowMapping/dict objects (bracket notation)
    def get_val(obj, key, default=None):
        if hasattr(obj, key):
            return getattr(obj, key)
        try:
            return obj[key]
        except (KeyError, TypeError):
            return default

    return {
        "id": get_val(cert, "id"),
        "tank_id": get_val(cert, "tank_id"),
        "tank_number": get_val(cert, "tank_number") or "",
        "certificate_number": get_val(cert, "certificate_number"),
        "insp_2_5y_date": safe_serialize_date(get_val(cert, "insp_2_5y_date")),
        "next_insp_date": safe_serialize_date(get_val(cert, "next_insp_date")),
        "year_of_manufacturing": get_val(cert, "year_of_manufacturing") or "",
        "inspection_agency": get_val(cert, "inspection_agency") or "",
        "status": 0 if get_val(cert, "status", 1) == 0 else 1,
        "remarks": get_val(cert, "remarks") or "",
        "archives": get_val(cert, "archives", 0) or 0,
        # certificate file URLs (6 slots)
        "initial_certificate_path": to_url(get_val(cert, "initial_certificate_path")),
        "initial_certificate_name": get_val(cert, "initial_certificate_name"),
        "certificate1_path": to_url(get_val(cert, "certificate1_path")),
        "certificate1_name": get_val(cert, "certificate1_name"),
        "certificate2_path": to_url(get_val(cert, "certificate2_path")),
        "certificate2_name": get_val(cert, "certificate2_name"),
        "certificate3_path": to_url(get_val(cert, "certificate3_path")),
        "certificate3_name": get_val(cert, "certificate3_name"),
        "certificate4_path": to_url(get_val(cert, "certificate4_path")),
        "certificate4_name": get_val(cert, "certificate4_name"),
        "certificate5_path": to_url(get_val(cert, "certificate5_path")),
        "certificate5_name": get_val(cert, "certificate5_name"),
        "created_by": get_val(cert, "created_by"),
        "updated_by": get_val(cert, "updated_by"),
        "created_at": get_val(cert, "created_at").isoformat() if get_val(cert, "created_at") else None,
        "updated_at": get_val(cert, "updated_at").isoformat() if get_val(cert, "updated_at") else None,
    }


def save_certificate(upload_file: Optional[UploadFile], label: str, image_type: str, tank_number: str) -> tuple:
    """Validate, save a certificate file and return (path, filename). Returns (None, None) if no file."""
    if not upload_file or not upload_file.filename:
        return None, None
    validate_certificate(upload_file, label)

    # Filename must contain the numeric part of the tank number
    numeric_tank = re.sub(r'\D', '', tank_number)
    numeric_file = re.sub(r'\D', '', upload_file.filename)
    if numeric_tank and numeric_tank not in numeric_file:
        raise HTTPException(
            status_code=400,
            detail=f"{label}: Filename mismatch. The file name must contain the numeric part '{numeric_tank}' from tank number '{tank_number}'."
        )

    path = save_uploaded_file(
        upload_file=upload_file,
        tank_number=tank_number,
        image_type=image_type,
        upload_root=UPLOAD_ROOT,
    )
    return path, upload_file.filename


def _shift_and_archive(cert, new_path: str, new_name: str):
    """
    Cascade shift certificates: slots 1→2→3→4→5→archive.
    The new file occupies slot 1.  Whatever was in slot 5 is considered archived
    and the `archives` counter is incremented.

    cert5 displaced  → archives += 1  (file is deleted from storage)
    cert4 → cert5
    cert3 → cert4
    cert2 → cert3
    cert1 → cert2
    new   → cert1
    """
    # If cert5 has a file, it gets pushed to archive (delete physically, increment counter)
    if cert.certificate5_path:
        delete_file_if_exists(UPLOAD_ROOT, cert.certificate5_path)
        cert.archives = int(cert.archives or 0) + 1

    # Shift 4 → 5
    cert.certificate5_path = cert.certificate4_path
    cert.certificate5_name = cert.certificate4_name

    # Shift 3 → 4
    cert.certificate4_path = cert.certificate3_path
    cert.certificate4_name = cert.certificate3_name

    # Shift 2 → 3
    cert.certificate3_path = cert.certificate2_path
    cert.certificate3_name = cert.certificate2_name

    # Shift 1 → 2
    cert.certificate2_path = cert.certificate1_path
    cert.certificate2_name = cert.certificate1_name

    # New file → slot 1
    cert.certificate1_path = new_path
    cert.certificate1_name = new_name


# -------- LIST ALL --------
@router.get("/")
def get_all_certificates(db: Session = Depends(get_db)):
    try:
        results = db.execute(text("CALL sp_GetAllCertificates()")).mappings().fetchall()
        return [serialize_certificate(r) for r in results]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error retrieving certificates: {str(e)}")

# -------- EXPORT TO EXCEL --------
@router.get("/export-to-excel")
def export_certificates_to_excel(
    nearing: Optional[bool] = False,
    db: Session = Depends(get_db)
):
    results = db.execute(text("CALL sp_GetAllCertificates()")).mappings().fetchall()
    
    # Convert mappings to objects for serialize helper
    cert_objects = [type('obj', (object,), dict(r))() for r in results]
    
    if nearing:
        filtered = []
        today = date_type.today()
        # Today's reference: total months from year 0
        current_total_months = today.year * 12 + (today.month - 1)
        
        for cert in cert_objects:
            if not getattr(cert, 'next_insp_date', None):
                continue
            try:
                # Format is YYYY/MM
                y_str, m_str = cert.next_insp_date.split('/')
                y, m = int(y_str), int(m_str)
                cert_total_months = y * 12 + (m - 1)
                
                diff = cert_total_months - current_total_months
                # Nearing if it's within 6 months (including already expired/past)
                if diff <= 6:
                    filtered.append(cert)
            except Exception:
                continue
        cert_objects = filtered

    if not results:
        raise HTTPException(status_code=404, detail="No certificate records found to export")

    wb = Workbook()
    ws = wb.active
    ws.title = "Certificates"

    headers = [
        "ID", "Tank Number", "Certificate Number", "Inspection Agency",
        "2.5Y Insp Date", "Next Insp Date", "Status", "Remarks", "Archives"
    ]

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row_num, cert in enumerate(cert_objects, 2):
        ws.cell(row=row_num, column=1, value=cert.id)
        ws.cell(row=row_num, column=2, value=getattr(cert, 'tank_number', ''))
        ws.cell(row=row_num, column=3, value=cert.certificate_number)
        ws.cell(row=row_num, column=4, value=cert.inspection_agency)
        ws.cell(row=row_num, column=5, value=safe_serialize_date(cert.insp_2_5y_date))
        ws.cell(row=row_num, column=6, value=safe_serialize_date(cert.next_insp_date))
        ws.cell(row=row_num, column=7, value="Active" if getattr(cert, "status", 1) == 1 else "Inactive")
        ws.cell(row=row_num, column=8, value=cert.remarks)
        ws.cell(row=row_num, column=9, value=cert.archives or 0)

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    prefix = "nearing_6m_" if nearing else ""
    filename = f"{prefix}certificates_export_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# -------- READ BY TANK ID --------
@router.get("/tank/{tank_id}")
def get_tank_certificates_by_tank(tank_id: int, db: Session = Depends(get_db)):
    try:
        results = db.execute(text("CALL sp_GetCertificatesByTankId(:tid)"), {"tid": tank_id}).mappings().fetchall()
        return [serialize_certificate(r) for r in results]
    except OperationalError as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Database schema error.")
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail="Error retrieving certificates")


# -------- READ BY ID --------
@router.get("/{cert_id}")
def get_tank_certificate_by_id(cert_id: int, db: Session = Depends(get_db)):
    result = db.execute(text("CALL sp_GetCertificateById(:id)"), {"id": cert_id}).mappings().first()
    if not result:
        raise HTTPException(status_code=404, detail="Tank certificate not found")
    return serialize_certificate(result)


# -------- CREATE --------
@router.post("/")
@router.post("/{path_tank_id}")
def create_tank_certificate(
    path_tank_id: Optional[int] = None,
    tank_id: Optional[int] = Form(None),
    certificate_number: Optional[str] = Form(None),
    insp_2_5y_date: Optional[str] = Form(None),
    next_insp_date: Optional[str] = Form(None),
    inspection_agency_id: Optional[int] = Form(None),
    # 6 certificate uploads
    initial_certificate: Optional[UploadFile] = File(None),
    certificate1: Optional[UploadFile] = File(None),
    certificate2: Optional[UploadFile] = File(None),
    certificate3: Optional[UploadFile] = File(None),
    certificate4: Optional[UploadFile] = File(None),
    certificate5: Optional[UploadFile] = File(None),
    remarks: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
):
    if path_tank_id is not None:
        tank_id = path_tank_id
    if tank_id is None:
        raise HTTPException(status_code=400, detail="tank_id is required")

    # Use SP to get tank record
    tank_record = db.execute(text("CALL sp_GetTankById(:tid)"), {"tid": tank_id}).mappings().first()
    if not tank_record:
        raise HTTPException(status_code=404, detail="Tank not found")

    existing_cert = db.execute(text("CALL sp_GetCertificatesByTankId(:tid)"), {"tid": tank_id}).mappings().first()
    if existing_cert:
        raise HTTPException(
            status_code=400,
            detail="A certificate already exists for this tank. Please edit the existing one."
        )

    tank_number = tank_record.tank_number

    # Auto-calculate next_insp_date if missing but insp_2_5y_date exists
    if not next_insp_date and insp_2_5y_date:
        next_insp_date = calculate_next_inspection_date(insp_2_5y_date)

    if not certificate_number or not certificate_number.strip():
        ts = datetime.now().strftime("%Y%m%d%H%M%S")
        certificate_number = f"CERT-{tank_number}-{ts}"

    emp_id = get_emp_id_from_token(authorization)

    # Look up inspection agency name
    inspection_agency_name = None
    if inspection_agency_id is not None:
        try:
            from app.models.inspection_agency_master_model import InspectionAgencyMaster
            agency = db.query(InspectionAgencyMaster).filter_by(id=inspection_agency_id).first()
            if agency:
                inspection_agency_name = agency.agency_name
        except Exception:
            pass

    # Fetch year_of_manufacturing
    year_of_manufacturing = tank_record.get("date_mfg") if tank_record else None

    # Save certificates (on create, all slots including initial are accepted)
    try:
        ic_path, ic_name = save_certificate(initial_certificate, "Initial Certificate", "cert_initial", tank_number)
        p1_path, p1_name = save_certificate(certificate1, "Certificate 1", "cert_certificate1", tank_number)
        p2_path, p2_name = save_certificate(certificate2, "Certificate 2", "cert_certificate2", tank_number)
        p3_path, p3_name = save_certificate(certificate3, "Certificate 3", "cert_certificate3", tank_number)
        p4_path, p4_name = save_certificate(certificate4, "Certificate 4", "cert_certificate4", tank_number)
        p5_path, p5_name = save_certificate(certificate5, "Certificate 5", "cert_certificate5", tank_number)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"File upload failed: {str(e)}")

    try:
        # Use SP for creation
        res = db.execute(
            text("CALL sp_UpsertCertificate(:id, :tid, :tnum, :cnum, :insp, :next, :agency, :ymfg, :rem, :stat, :arch, :icp, :icn, :c1p, :c1n, :c2p, :c2n, :c3p, :c3n, :c4p, :c4n, :c5p, :c5n, :eid)"),
            {
                "id": 0, "tid": tank_id, "tnum": tank_number, "cnum": clean_form_data(certificate_number),
                "insp": _normalize_date_str(insp_2_5y_date), "next": _normalize_date_str(next_insp_date),
                "agency": inspection_agency_name, "ymfg": year_of_manufacturing, "rem": clean_form_data(remarks),
                "stat": 1, "arch": 0, "icp": ic_path, "icn": ic_name,
                "c1p": p1_path, "c1n": p1_name, "c2p": p2_path, "c2n": p2_name,
                "c3p": p3_path, "c3n": p3_name, "c4p": p4_path, "c4n": p4_name,
                "c5p": p5_path, "c5n": p5_name, "eid": emp_id
            }
        ).mappings().first()
        
        db.commit()
        # Fetch the newly created cert
        cert_row = db.execute(text("CALL sp_GetCertificateById(:id)"), {"id": res["id"]}).mappings().first()
    except Exception as e:
        db.rollback()
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=400, detail=f"Database Insertion Failed: {str(e)}")

    # Sync next inspection date to tank_inspection_details
    try:
        db.execute(
            text("CALL sp_SyncNextInspectionDate(:tnum, :next_date)"),
            {"next_date": _normalize_date_str(next_insp_date), "tnum": tank_number}
        )
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to sync pi_next_inspection_date: {str(e)}")

    return {"message": "Tank certificate added successfully", "data": serialize_certificate(cert_row)}


# -------- UPDATE --------
@router.put("/{cert_id}")
def update_tank_certificate(
    cert_id: int,
    certificate_number: Optional[str] = Form(None),
    insp_2_5y_date: Optional[str] = Form(None),
    next_insp_date: Optional[str] = Form(None),
    inspection_agency_id: Optional[int] = Form(None),
    status: Optional[int] = Form(None),
    # certificate1 is the ONLY replaceable slot — triggers the cascade shift
    initial_certificate: Optional[UploadFile] = File(None),
    remove_initial_certificate: Optional[bool] = Form(False),
    certificate1: Optional[UploadFile] = File(None),
    remove_certificate1: Optional[bool] = Form(False),
    certificate2: Optional[UploadFile] = File(None),
    remove_certificate2: Optional[bool] = Form(False),
    certificate3: Optional[UploadFile] = File(None),
    remove_certificate3: Optional[bool] = Form(False),
    certificate4: Optional[UploadFile] = File(None),
    remove_certificate4: Optional[bool] = Form(False),
    certificate5: Optional[UploadFile] = File(None),
    remove_certificate5: Optional[bool] = Form(False),
    remarks: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
):
    """
    Update a tank certificate record.

    Rules:
    - initial_certificate: immutable after first upload — cannot be replaced or removed via API.
    - certificate1: uploading a new file triggers cascade shift:
        existing cert1 → cert2, cert2 → cert3, cert3 → cert4, cert4 → cert5,
        cert5 → archived (physically deleted, archives counter +1).
    - certificate2–5 are read-only; they are populated automatically by the shift.
    """
    result = db.execute(text("CALL sp_GetCertificateById(:id)"), {"id": cert_id}).mappings().first()
    if not result:
        raise HTTPException(status_code=404, detail="Tank certificate not found")
    
    # Create mutable object for updates
    cert = type('obj', (object,), dict(result))()

    emp_id = get_emp_id_from_token(authorization)

    tank_record = db.execute(text("CALL sp_GetTankById(:tid)"), {"tid": cert.tank_id}).mappings().first()
    tank_number = tank_record.get("tank_number") if tank_record else "UNKNOWN"

    if certificate_number is not None:
        cert.certificate_number = clean_form_data(certificate_number)
    if insp_2_5y_date is not None:
        n = _normalize_date_str(insp_2_5y_date)
        if n:
            cert.insp_2_5y_date = n
            # If next_insp_date not provided, auto-update it
            if next_insp_date is None:
                cert.next_insp_date = calculate_next_inspection_date(n)
    if next_insp_date is not None:
        n = _normalize_date_str(next_insp_date)
        if n:
            cert.next_insp_date = n
    if status is not None:
        cert.status = 1 if int(status) == 1 else 0
    if inspection_agency_id is not None:
        try:
            from app.models.inspection_agency_master_model import InspectionAgencyMaster
            agency = db.query(InspectionAgencyMaster).filter_by(id=inspection_agency_id).first()
            if not agency:
                raise HTTPException(status_code=400, detail=f"inspection_agency_id '{inspection_agency_id}' is not valid")
            cert.inspection_agency = agency.agency_name
        except HTTPException:
            raise
        except Exception:
            pass

    if remarks is not None:
        cert.remarks = clean_form_data(remarks)

    try:
        # Handle initial_certificate
        if remove_initial_certificate:
            if initial_certificate and initial_certificate.filename:
                new_path, new_name = save_certificate(initial_certificate, "Initial Certificate", "cert_initial", tank_number)
                if cert.initial_certificate_path: delete_file_if_exists(UPLOAD_ROOT, cert.initial_certificate_path)
                cert.initial_certificate_path, cert.initial_certificate_name = new_path, new_name
            else:
                if cert.initial_certificate_path: delete_file_if_exists(UPLOAD_ROOT, cert.initial_certificate_path)
                cert.initial_certificate_path, cert.initial_certificate_name = None, None
        elif initial_certificate and initial_certificate.filename:
            new_path, new_name = save_certificate(initial_certificate, "Initial Certificate", "cert_initial", tank_number)
            if cert.initial_certificate_path: delete_file_if_exists(UPLOAD_ROOT, cert.initial_certificate_path)
            cert.initial_certificate_path, cert.initial_certificate_name = new_path, new_name

        # Handle certificate1 replacement with cascade shift
        if remove_certificate1:
            if certificate1 and certificate1.filename:
                # Replace WITHOUT cascade shift
                new_path, new_name = save_certificate(certificate1, "Certificate 1", "cert_certificate1", tank_number)
                if cert.certificate1_path:
                    delete_file_if_exists(UPLOAD_ROOT, cert.certificate1_path)
                cert.certificate1_path = new_path
                cert.certificate1_name = new_name
            else:
                # Remove WITHOUT cascade shift
                if cert.certificate1_path:
                    delete_file_if_exists(UPLOAD_ROOT, cert.certificate1_path)
                cert.certificate1_path = None
                cert.certificate1_name = None
        elif certificate1 and certificate1.filename:
            # Normal replacement WITH cascade shift
            new_path, new_name = save_certificate(certificate1, "Certificate 1", "cert_certificate1", tank_number)
            _shift_and_archive(cert, new_path, new_name)

        # Handle certificate2..5
        for idx, file_obj, remove_flag in [
            (2, certificate2, remove_certificate2),
            (3, certificate3, remove_certificate3),
            (4, certificate4, remove_certificate4),
            (5, certificate5, remove_certificate5),
        ]:
            path_attr = f"certificate{idx}_path"
            name_attr = f"certificate{idx}_name"

            if remove_flag:
                if file_obj and file_obj.filename:
                    n_path, n_name = save_certificate(file_obj, f"Certificate {idx}", f"cert_certificate{idx}", tank_number)
                    if getattr(cert, path_attr): delete_file_if_exists(UPLOAD_ROOT, getattr(cert, path_attr))
                    setattr(cert, path_attr, n_path)
                    setattr(cert, name_attr, n_name)
                else:
                    if getattr(cert, path_attr): delete_file_if_exists(UPLOAD_ROOT, getattr(cert, path_attr))
                    setattr(cert, path_attr, None)
                    setattr(cert, name_attr, None)
            elif file_obj and file_obj.filename:
                n_path, n_name = save_certificate(file_obj, f"Certificate {idx}", f"cert_certificate{idx}", tank_number)
                if getattr(cert, path_attr): delete_file_if_exists(UPLOAD_ROOT, getattr(cert, path_attr))
                setattr(cert, path_attr, n_path)
                setattr(cert, name_attr, n_name)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"File update failed: {str(e)}")

    cert.updated_by = emp_id

    try:
        db.execute(
            text("CALL sp_UpsertCertificate(:id, :tid, :tnum, :cnum, :insp, :next, :agency, :ymfg, :rem, :stat, :arch, :icp, :icn, :c1p, :c1n, :c2p, :c2n, :c3p, :c3n, :c4p, :c4n, :c5p, :c5n, :eid)"),
            {
                "id": cert_id, "tid": cert.tank_id, "tnum": tank_number, "cnum": cert.certificate_number,
                "insp": cert.insp_2_5y_date, "next": cert.next_insp_date,
                "agency": cert.inspection_agency, "ymfg": cert.year_of_manufacturing, "rem": cert.remarks,
                "stat": cert.status, "arch": cert.archives, "icp": cert.initial_certificate_path, "icn": cert.initial_certificate_name,
                "c1p": cert.certificate1_path, "c1n": cert.certificate1_name, "c2p": cert.certificate2_path, "c2n": cert.certificate2_name,
                "c3p": cert.certificate3_path, "c3n": cert.certificate3_name, "c4p": cert.certificate4_path, "c4n": cert.certificate4_name,
                "c5p": cert.certificate5_path, "c5n": cert.certificate5_name, "eid": emp_id
            }
        )
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(traceback.format_exc())
        raise HTTPException(status_code=400, detail=f"Database Update Failed: {str(e)}")

    try:
        db.execute(
            text("CALL sp_SyncNextInspectionDate(:tnum, :next_date)"),
            {"next_date": cert.next_insp_date, "tank_num": tank_number}
        )
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to sync pi_next_inspection_date on update: {str(e)}")

    return {"message": "Tank certificate updated successfully", "data": serialize_certificate(cert)}


# -------- DELETE --------
@router.delete("/{cert_id}")
def delete_tank_certificate(
    cert_id: int,
    db: Session = Depends(get_db),
    authorization: Optional[str] = Header(None),
):
    result = db.execute(text("CALL sp_GetCertificateById(:id)"), {"id": cert_id}).mappings().first()
    if not result:
        raise HTTPException(status_code=404, detail="Tank certificate not found")
    
    cert = type('obj', (object,), dict(result))()

    for field in ["initial_certificate_path", "certificate1_path", "certificate2_path", "certificate3_path", "certificate4_path", "certificate5_path"]:
        p = getattr(cert, field, None)
        if p:
            delete_file_if_exists(UPLOAD_ROOT, p)

    db.execute(text("CALL sp_DeleteCertificate(:id)"), {"id": cert_id})
    db.commit()
    return {"message": "Tank certificate deleted successfully"}
